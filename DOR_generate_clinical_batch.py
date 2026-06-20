# DOR_generate_clinical_batch.py
import argparse
import copy
import os
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

ALLOWED_NATURES = {"Sympt", "Signe"}
ALLOWED_SCORES = {50, 100, 150}

ALIASES = {
    "CodeMaladie": ["CodeMaladie", "code maladie", "code_maladie", "DiseaseCode", "Code Disease"],
    "NatureLien": ["NatureLien", "nature lien", "nature_lien", "Nature", "Type", "Lien"],
    "CodeElement": ["CodeElement", "code element", "code_element", "ElementCode", "Code"],
    "Score": ["Score", "score", "Poids", "Weight"],
}

DEFAULT_BATCH_SHEETS = {
    "tabel2": ["BATCH_Tabel2", "Tabel2", "Tabel2_ROWS", "Tabel2_rows", "Rows"],
    "sympt": ["BATCH_Symptomes", "BATCH_Symptomes_NEW", "Symptomes_NEW", "Symptomes1", "Sympt"],
    "signe": ["BATCH_Signes", "BATCH_Signe_NEW", "Signe_NEW", "Signe1", "Signes"],
    "maladies": ["BATCH_Maladies", "BATCH_Maladies_UPDATE", "Maladies_UPDATE", "Maladies"],
}


def norm_header(value):
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]", "", str(value).strip().lower())


def code4(value):
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""
    if re.fullmatch(r"\d+", s):
        return s.zfill(4)
    return s


def code_element(value):
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""
    if re.fullmatch(r"\d+", s):
        return s.zfill(4)
    return s


def score_value(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return int(value)
    s = str(value).strip()
    if not s:
        return None
    return int(float(s))


def find_header_row(ws, max_scan=15):
    for row in range(1, min(ws.max_row, max_scan) + 1):
        headers = [norm_header(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
        hits = 0
        for aliases in ALIASES.values():
            alias_norms = {norm_header(a) for a in aliases}
            if any(h in alias_norms for h in headers):
                hits += 1
        if hits >= 2:
            return row
    return 1


def map_headers(ws, header_row):
    raw = {col: ws.cell(header_row, col).value for col in range(1, ws.max_column + 1)}
    normalized = {col: norm_header(value) for col, value in raw.items()}
    mapped = {}
    for canonical, aliases in ALIASES.items():
        alias_norms = {norm_header(a) for a in aliases}
        for col, header in normalized.items():
            if header in alias_norms:
                mapped[canonical] = col
                break
    return raw, mapped


def require_columns(ws, mapped, required, label):
    missing = [c for c in required if c not in mapped]
    if missing:
        raise ValueError(f"{label}: missing required columns: {missing}; sheet={ws.title}")


def get_sheet(wb, candidates, required=False, label="sheet"):
    lower = {s.lower(): s for s in wb.sheetnames}
    for candidate in candidates:
        if candidate.lower() in lower:
            return wb[lower[candidate.lower()]]
    if required:
        raise ValueError(f"Missing {label}. Tried sheet names: {candidates}. Found: {wb.sheetnames}")
    return None


def row_is_empty(ws, row):
    return all(ws.cell(row, col).value in (None, "") for col in range(1, ws.max_column + 1))


def first_empty_row(ws, header_row=1):
    row = ws.max_row + 1
    while row > header_row + 1 and row_is_empty(ws, row - 1):
        row -= 1
    return row


def existing_tabel2_keys(ws, mapped, header_row):
    keys = set()
    for row in range(header_row + 1, ws.max_row + 1):
        cm = code4(ws.cell(row, mapped["CodeMaladie"]).value)
        nl = str(ws.cell(row, mapped["NatureLien"]).value or "").strip()
        ce = code_element(ws.cell(row, mapped["CodeElement"]).value)
        if cm and nl and ce:
            keys.add((cm, nl, ce))
    return keys


def existing_catalog_codes(ws):
    header_row = find_header_row(ws)
    _, mapped = map_headers(ws, header_row)
    code_col = mapped.get("CodeElement") or mapped.get("CodeMaladie")
    codes = set()
    if code_col:
        for row in range(header_row + 1, ws.max_row + 1):
            value = ws.cell(row, code_col).value
            if value not in (None, ""):
                codes.add(code_element(value))
    return codes


def copy_row_values_and_style(src_ws, src_row, dst_ws, dst_row, max_col=None):
    max_col = max_col or src_ws.max_column
    for col in range(1, max_col + 1):
        src = src_ws.cell(src_row, col)
        dst = dst_ws.cell(dst_row, col)
        dst.value = src.value
        if src.has_style:
            dst._style = copy.copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy.copy(src.alignment)
        if src.fill:
            dst.fill = copy.copy(src.fill)
        if src.font:
            dst.font = copy.copy(src.font)
        if src.border:
            dst.border = copy.copy(src.border)
        if src.protection:
            dst.protection = copy.copy(src.protection)
        if src.comment:
            dst.comment = copy.copy(src.comment)


def append_rows_matching_headers(src_ws, dst_ws, dst_header_row, source_label, required_columns=None, dedupe_key=None):
    required_columns = required_columns or []
    src_header_row = find_header_row(src_ws)
    src_headers, src_map = map_headers(src_ws, src_header_row)
    dst_headers, dst_map = map_headers(dst_ws, dst_header_row)
    require_columns(src_ws, src_map, required_columns, source_label)

    dst_header_norm_to_col = {norm_header(v): c for c, v in dst_headers.items() if v not in (None, "")}
    src_to_dst = {}
    for src_col, header in src_headers.items():
        n = norm_header(header)
        if n and n in dst_header_norm_to_col:
            src_to_dst[src_col] = dst_header_norm_to_col[n]

    if not src_to_dst:
        raise ValueError(f"{source_label}: no matching headers between {src_ws.title} and {dst_ws.title}")

    appended = 0
    skipped = 0
    existing_keys = set()
    if dedupe_key == "tabel2":
        existing_keys = existing_tabel2_keys(dst_ws, dst_map, dst_header_row)
    elif dedupe_key == "catalog_code":
        existing_keys = existing_catalog_codes(dst_ws)

    template_row = max(dst_header_row + 1, first_empty_row(dst_ws, dst_header_row) - 1)

    for src_row in range(src_header_row + 1, src_ws.max_row + 1):
        if row_is_empty(src_ws, src_row):
            continue
        if dedupe_key == "tabel2":
            key = (
                code4(src_ws.cell(src_row, src_map["CodeMaladie"]).value),
                str(src_ws.cell(src_row, src_map["NatureLien"]).value or "").strip(),
                code_element(src_ws.cell(src_row, src_map["CodeElement"]).value),
            )
            if key in existing_keys:
                skipped += 1
                continue
            existing_keys.add(key)
        elif dedupe_key == "catalog_code":
            code_col = src_map.get("CodeElement") or src_map.get("CodeMaladie")
            key = code_element(src_ws.cell(src_row, code_col).value) if code_col else None
            if key and key in existing_keys:
                skipped += 1
                continue
            if key:
                existing_keys.add(key)

        dst_row = first_empty_row(dst_ws, dst_header_row)
        if template_row >= dst_header_row + 1:
            for col in range(1, dst_ws.max_column + 1):
                tmpl = dst_ws.cell(template_row, col)
                dst = dst_ws.cell(dst_row, col)
                if tmpl.has_style:
                    dst._style = copy.copy(tmpl._style)
                dst.number_format = tmpl.number_format
                dst.alignment = copy.copy(tmpl.alignment)
                dst.fill = copy.copy(tmpl.fill)
                dst.font = copy.copy(tmpl.font)
                dst.border = copy.copy(tmpl.border)
                dst.protection = copy.copy(tmpl.protection)
        for src_col, dst_col in src_to_dst.items():
            value = src_ws.cell(src_row, src_col).value
            header_norm = norm_header(dst_ws.cell(dst_header_row, dst_col).value)
            if header_norm in {norm_header("CodeMaladie"), norm_header("DiseaseCode")}:
                value = code4(value)
            if header_norm in {norm_header("CodeElement"), norm_header("ElementCode"), norm_header("Code")}:
                value = code_element(value)
            if header_norm == norm_header("Score") and value not in (None, ""):
                value = score_value(value)
            dst_ws.cell(dst_row, dst_col).value = value
        appended += 1

    return {"appended": appended, "skipped_duplicates": skipped}


def validate_tabel2(ws, min_code=None, max_code=None):
    header_row = find_header_row(ws)
    _, mapped = map_headers(ws, header_row)
    require_columns(ws, mapped, ["CodeMaladie", "NatureLien", "CodeElement", "Score"], "Tabel2 validation")

    counts = defaultdict(int)
    bad = []
    keys = set()
    dupes = []
    nature_counts = Counter()
    score_counts = Counter()

    for row in range(header_row + 1, ws.max_row + 1):
        if row_is_empty(ws, row):
            continue
        cm = code4(ws.cell(row, mapped["CodeMaladie"]).value)
        nl = str(ws.cell(row, mapped["NatureLien"]).value or "").strip()
        ce = code_element(ws.cell(row, mapped["CodeElement"]).value)
        sc = score_value(ws.cell(row, mapped["Score"]).value)
        if min_code and max_code and re.fullmatch(r"\d{4}", cm):
            if not (min_code <= cm <= max_code):
                continue
        key = (cm, nl, ce)
        if key in keys:
            dupes.append((row, key))
        keys.add(key)
        counts[cm] += 1
        nature_counts[nl] += 1
        score_counts[sc] += 1
        if nl not in ALLOWED_NATURES:
            bad.append((row, "NatureLien", nl))
        if sc not in ALLOWED_SCORES:
            bad.append((row, "Score", sc))
        if not cm or not ce:
            bad.append((row, "MissingCode", key))

    count_violations = [(cm, n) for cm, n in counts.items() if n < 10 or n > 15]
    return {
        "row_count": sum(counts.values()),
        "disease_count": len(counts),
        "nature_counts": dict(nature_counts),
        "score_counts": dict(score_counts),
        "count_violations": count_violations,
        "bad_values": bad,
        "duplicate_keys": dupes,
    }


def add_report_sheet(wb, report_name, items):
    if report_name in wb.sheetnames:
        del wb[report_name]
    ws = wb.create_sheet(report_name)
    ws.append(["Metric", "Value"])
    for k, v in items:
        ws.append([k, str(v)])
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 100
    return ws


def save_copy(wb, source_path, output_path):
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    if not Path(output_path).exists() or Path(output_path).stat().st_size == 0:
        raise IOError(f"Output was not created correctly: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Merge a clinically validated batch into complete MedDiag Excel workbooks.")
    parser.add_argument("--base-tabel2", required=True)
    parser.add_argument("--base-symptomes", required=True)
    parser.add_argument("--base-signe", required=True)
    parser.add_argument("--base-maladies", required=True)
    parser.add_argument("--batch", required=True, help="Workbook containing BATCH_Tabel2 and optional BATCH_Symptomes_NEW/BATCH_Signe_NEW/BATCH_Maladies_UPDATE sheets")
    parser.add_argument("--out-dir", required=True)
    parser.add_argument("--end-code", required=True, help="Example: 0638")
    parser.add_argument("--min-batch-code", default=None, help="Optional validation lower bound, example: 0588")
    parser.add_argument("--max-batch-code", default=None, help="Optional validation upper bound, example: 0638")
    args = parser.parse_args()

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    base_tabel2 = load_workbook(args.base_tabel2)
    base_sympt = load_workbook(args.base_symptomes)
    base_signe = load_workbook(args.base_signe)
    base_maladies = load_workbook(args.base_maladies)
    batch = load_workbook(args.batch, data_only=False)

    tabel2_ws = base_tabel2.active
    sympt_ws = base_sympt.active
    signe_ws = base_signe.active
    maladies_ws = base_maladies.active

    tabel2_header = find_header_row(tabel2_ws)
    sympt_header = find_header_row(sympt_ws)
    signe_header = find_header_row(signe_ws)
    maladies_header = find_header_row(maladies_ws)

    batch_tabel2_ws = get_sheet(batch, DEFAULT_BATCH_SHEETS["tabel2"], required=True, label="batch Tabel2 sheet")
    stats = []
    stats.append(("Tabel2 append", append_rows_matching_headers(batch_tabel2_ws, tabel2_ws, tabel2_header, "BATCH_Tabel2", ["CodeMaladie", "NatureLien", "CodeElement", "Score"], "tabel2")))

    batch_sympt_ws = get_sheet(batch, DEFAULT_BATCH_SHEETS["sympt"], required=False)
    if batch_sympt_ws:
        stats.append(("Symptomes append", append_rows_matching_headers(batch_sympt_ws, sympt_ws, sympt_header, "BATCH_Symptomes", [], "catalog_code")))
    else:
        stats.append(("Symptomes append", "no batch sheet found"))

    batch_signe_ws = get_sheet(batch, DEFAULT_BATCH_SHEETS["signe"], required=False)
    if batch_signe_ws:
        stats.append(("Signe append", append_rows_matching_headers(batch_signe_ws, signe_ws, signe_header, "BATCH_Signe", [], "catalog_code")))
    else:
        stats.append(("Signe append", "no batch sheet found"))

    batch_maladies_ws = get_sheet(batch, DEFAULT_BATCH_SHEETS["maladies"], required=False)
    if batch_maladies_ws:
        stats.append(("Maladies append", append_rows_matching_headers(batch_maladies_ws, maladies_ws, maladies_header, "BATCH_Maladies", [], "catalog_code")))
    else:
        stats.append(("Maladies append", "no batch sheet found"))

    validation = validate_tabel2(tabel2_ws, args.min_batch_code, args.max_batch_code)
    if validation["bad_values"] or validation["duplicate_keys"] or validation["count_violations"]:
        report_path = out_dir / f"DOR_VALIDATION_ERRORS_0001_{args.end_code}.txt"
        with open(report_path, "w", encoding="utf-8") as f:
            f.write(str(validation))
        raise ValueError(f"Validation failed. See {report_path}")

    report_items = stats + [("Validation", validation)]
    add_report_sheet(base_tabel2, "DOR_Generator_Report", report_items)

    out_tabel2 = out_dir / f"Tabel2_Nou_0001_{args.end_code}_CLINIC_FORMAL_VALIDATED.xlsx"
    out_sympt = out_dir / f"Symptomes1_UPDATED_0001_{args.end_code}.xlsx"
    out_signe = out_dir / f"Signe1_UPDATED_0001_{args.end_code}.xlsx"
    out_maladies = out_dir / f"Maladies_UPDATED_0001_{args.end_code}.xlsx"

    save_copy(base_tabel2, args.base_tabel2, out_tabel2)
    save_copy(base_sympt, args.base_symptomes, out_sympt)
    save_copy(base_signe, args.base_signe, out_signe)
    save_copy(base_maladies, args.base_maladies, out_maladies)

    print("OK: generated complete incremental workbooks")
    print(out_tabel2)
    print(out_sympt)
    print(out_signe)
    print(out_maladies)


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        sys.exit(1)
