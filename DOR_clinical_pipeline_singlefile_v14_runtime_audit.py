# DOR_clinical_pipeline_singlefile_v11_autosuggest.py
# Filename: DOR_clinical_pipeline_singlefile_v14_runtime_audit.py
#
# V13 = NlpRO pipeline with corrected neurological mappings and AutoSuggestMapping.
#
# Adds:
#   - Auto-suggestions for Unmatched utterances
#   - SuggestedCodeElement
#   - SuggestedNatureElement
#   - SuggestedCatalogName
#   - SuggestedElementStandard
#   - SuggestionMethod
#   - SuggestionConfidence
#   - SuggestionReason
#   - ValidationStatus
#
# Workflow:
#   1. Load/append transcript in NlpRO.
#   2. Run this script.
#   3. Review Unmatched suggestions.
#   4. Set ValidationStatus = accept for correct suggestions.
#   5. Rerun script: accepted suggestions are imported into Lexicon.
#
# Install:
#   pip install pandas openpyxl

import re
import json
import unicodedata
from pathlib import Path
from datetime import datetime
from difflib import SequenceMatcher

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
WORKBOOK = BASE_DIR / "ClinicalPipeline_RO_SINGLE_v14_RUNTIME_AUDIT.xlsx"

CONTEXT_COLUMNS = [
    "TriggerContext", "TemporalContext", "SeverityContext", "BodyRegion",
    "ClinicalIntent", "AggravatingFactor", "RelievingFactor",
]

SUGGESTION_COLUMNS = [
    "SuggestedExpression",
    "SuggestedElementStandard",
    "SuggestedCodeElement",
    "SuggestedNatureElement",
    "SuggestedCatalogName",
    "SuggestionMethod",
    "SuggestionConfidence",
    "SuggestionReason",
    "ValidationStatus",
]


def strip_diacritics(text):
    text = unicodedata.normalize("NFKD", str(text))
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def normalize_text(text):
    text = "" if text is None else str(text)
    text = text.lower().strip()
    text = text.replace("ş", "ș").replace("ţ", "ț")
    text = strip_diacritics(text)
    text = re.sub(r"[^a-z0-9\s]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def clean_code(v):
    if v is None:
        return ""
    s = str(v).strip()
    if not s or s.lower() in {"nan", "none", "null"}:
        return ""
    if re.fullmatch(r"\d+(\.0)?", s):
        return str(int(float(s))).zfill(4)
    return s.zfill(4) if s.isdigit() else s


def ensure_columns(df, cols):
    for c in cols:
        if c not in df.columns:
            df[c] = ""
    return df


def format_workbook(path):
    wb = load_workbook(path)
    colors = {
        "Lexicon": "0F766E",
        "Dialogues": "1D4ED8",
        "Annotations": "7C3AED",
        "PatientVectors": "0891B2",
        "MedDiagInput": "0369A1",
        "ClinicalSynthesis": "0E7490",
        "AutoSuggestions": "B45309",
        "Unmatched": "F97316",
        "Summary": "475569",
        "LoaderLog": "64748B",
        "MappingAudit_V13": "7C2D12",
        "LexiconRuntimeAudit": "164E63",
        "LexiconDialogMatchAudit": "155E75",
    }
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        fill = PatternFill("solid", fgColor=colors.get(ws.title, "0F766E"))
        for cell in ws[1]:
            cell.fill = fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col_idx, col in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in col[:80]:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(val), 70))
            ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(max_len + 2, 80))
    wb.save(path)


# --------------------------
# Matching / NLP core
# --------------------------

def detect_clinical_polarity(utterance_norm, expr_norm, lexicon_polarity="prezent"):
    u = normalize_text(utterance_norm)
    e = normalize_text(expr_norm)
    lex_pol = normalize_text(lexicon_polarity or "prezent")

    if lex_pol in {"absent", "negativ"}:
        return "absent", "lexicon_absent", "Lexicon polarity is absent."

    symptom_negation_phrases = [
        "nu mai dorm", "nu dorm", "nu pot respira", "nu pot sa respir",
        "nu pot să respir", "nu am aer", "nu mi ajunge aerul",
        "nu imi ajunge aerul", "nu mai am aer", "nu merg drept",
    ]
    for p in symptom_negation_phrases:
        pn = normalize_text(p)
        if pn and pn in u:
            return "prezent", pn, "Negation is part of the positive symptom expression."

    pos = u.find(e)
    local = u[max(0, pos - 45):pos] + " " + e if pos >= 0 else u

    resolved_patterns = [
        r"\bnu mai am\b", r"\bnu mai prezint\b", r"\bnu mai simt\b",
        r"\bmi a trecut\b", r"\ba trecut\b", r"\bs a oprit\b",
    ]
    absent_patterns = [
        r"\bnu am\b", r"\bn am\b", r"\bnu prezint\b",
        r"\bfara\b", r"\bfără\b", r"\bnici\b", r"\bniciun\b", r"\bnicio\b",
    ]

    for pat in resolved_patterns:
        if re.search(pat, local):
            return "rezolvat", pat, "Resolved cue found before/near matched expression."

    for pat in absent_patterns:
        if re.search(pat, local):
            return "absent", pat, "Negation cue found before/near matched expression."

    return "prezent", "", "Default present."


def detect_clinical_temporality(utterance_norm):
    u = normalize_text(utterance_norm)
    patterns = [
        (r"\b(brusc|dintr odata|dintr odată|din senin)\b", "sudden", "sudden_onset", ""),
        (r"\b(de azi|azi|de dimineata|de dimineață|de aseara|de aseară|azi noapte|azi-noapte)\b", "acute_recent", "acute", "hours_days"),
        (r"\b(de ieri|ieri)\b", "acute_recent", "acute", "days"),
        (r"\b(de cateva zile|de câteva zile|de [0-9]+ zile|de vreo [0-9]+ zile)\b", "acute_recent", "subacute", "days"),
        (r"\b(de o saptamana|de o săptămână|de [0-9]+ saptamani|de [0-9]+ săptămâni|de cateva saptamani|de câteva săptămâni)\b", "subacute", "subacute", "weeks"),
        (r"\b(de o luna|de o lună|de [0-9]+ luni|de cateva luni|de câteva luni)\b", "chronic", "gradual_or_chronic", "months"),
        (r"\b(de ani|de [0-9]+ ani|din copilarie|din copilărie)\b", "chronic_long", "chronic", "years"),
        (r"\b(de multa vreme|de multă vreme|de mult timp|de ceva vreme)\b", "chronic", "gradual_or_chronic", "months_or_longer"),
        (r"\b(uneori|din cand in cand|din când în când|vine si pleaca|vine și pleacă|intermitent)\b", "intermittent", "episodic", "intermittent"),
        (r"\b(noaptea|nopti|nopți|nocturn)\b", "nocturnal", "", "night"),
        (r"\b(progresiv|tot mai|din ce in ce|din ce în ce)\b", "progressive", "progressive", ""),
    ]
    for pat, temporal, onset, duration in patterns:
        m = re.search(pat, u)
        if m:
            return temporal, m.group(0), onset, duration
    return "", "", "", ""


def detect_clinical_severity(utterance_norm, lexicon_intensity=""):
    u = normalize_text(utterance_norm)
    lex_i = normalize_text(lexicon_intensity or "")

    severe_patterns = [
        r"\b(foarte tare|cumplit|cumplite|groaznic|groaznice|insuportabil|nu mai rezist|de moarte|ma rupe|mă rupe|ma omoara|mă omoară|sever|puternic|puternice|grav|doboratoare|doborâtoare)\b",
        r"\b(nu pot|abia pot|aproape lesin|aproape leșin)\b",
    ]
    mild_patterns = [r"\b(putin|puțin|usoara|ușoară|usor|ușor|jena|jenă|discret|slab)\b"]
    moderate_patterns = [r"\b(deranjant|ma deranjeaza|mă deranjează|ma supara|mă supără|moderat)\b"]

    for pat in severe_patterns:
        m = re.search(pat, u)
        if m:
            return "severe", m.group(0), 3, "Severe cue found in utterance."
    for pat in mild_patterns:
        m = re.search(pat, u)
        if m:
            return "mild", m.group(0), 1, "Mild cue found in utterance."
    for pat in moderate_patterns:
        m = re.search(pat, u)
        if m:
            return "moderate", m.group(0), 2, "Moderate cue found in utterance."

    if lex_i in {"puternic", "severe", "forte"}:
        return "severe", "lexicon_intensity", 3, "Severity inferred from lexicon intensity."
    if lex_i in {"slab", "mild", "faible"}:
        return "mild", "lexicon_intensity", 1, "Severity inferred from lexicon intensity."
    if lex_i in {"moderat", "moderate"}:
        return "moderate", "lexicon_intensity", 2, "Severity inferred from lexicon intensity."
    return "", "", "", ""


# --------------------------
# AutoSuggestMapping
# --------------------------

RULE_SUGGESTIONS = [
    {
        "patterns": ["batut cu maiul", "bătut cu maiul", "ca batut", "ca bătut", "sunt praf", "sunt terminat", "nu mai am putere", "nu mai am vlaga", "nu mai am vlagă", "slabiciune", "slăbiciune", "doboratoare", "doborâtoare", "organismul nu ma mai ajuta", "organismul nu mă mai ajută"],
        "expr": "",
        "code": "0014", "nature": "Sympt", "catalog": "FATIGUE", "standard": "Oboseală / astenie",
        "domain": "astenie", "confidence": 0.92, "reason": "Astenie/fatigue popular expression."
    },
    {
        "patterns": ["nu am aer", "nu mai am aer", "nu mi ajunge aerul", "nu imi ajunge aerul", "parca nu mi ajunge aerul", "mă sufoc", "ma sufoc", "gafai", "gâfâi", "fara aer", "fără aer"],
        "expr": "",
        "code": "0006", "nature": "Sympt", "catalog": "DYSPNEA", "standard": "Dispnee",
        "domain": "respirator", "confidence": 0.95, "reason": "Dyspnea / lack of air expression."
    },
    {
        "patterns": ["nu mai dorm", "nu dorm", "dorm prost", "ma scol in mijlocul noptii", "mă scol în mijlocul nopții", "ma trezesc noaptea", "mă trezesc noaptea", "nopțile", "noptile", "nu pun geana"],
        "expr": "",
        "code": "0498", "nature": "Sympt", "catalog": "INSOMNIA", "standard": "Insomnie",
        "domain": "somn", "confidence": 0.95, "reason": "Insomnia / sleep disruption expression."
    },
    {
        "patterns": ["grijile", "ma coplesesc", "mă copleșesc", "coplesesc", "neliniste", "neliniște", "anxios", "frica", "frică", "ma sperii", "mă sperii"],
        "expr": "",
        "code": "0045", "nature": "Sympt", "catalog": "ANXIETY", "standard": "Anxietate",
        "domain": "psihic", "confidence": 0.90, "reason": "Anxiety / worries expression."
    },
    {
        "patterns": ["ma doare capul", "mă doare capul", "ma tine capul", "mă ține capul", "durere de cap", "dureri de cap", "imi crapa capul", "îmi crapă capul"],
        "expr": "",
        "code": "0023", "nature": "Sympt", "catalog": "HEADACHE", "standard": "Cefalee",
        "domain": "durere", "confidence": 0.92, "reason": "Headache expression."
    },
    {
        "patterns": ["febr", "temperatura", "temperatură", "arde fruntea", "mi a ars fruntea", "îmi arde fruntea"],
        "expr": "",
        "code": "0022", "nature": "Sympt", "catalog": "FEVER", "standard": "Febră",
        "domain": "general", "confidence": 0.90, "reason": "Fever expression."
    },
    {
        "patterns": ["tusesc", "tușesc", "tuse", "ma rupe tusea", "mă rupe tusea"],
        "expr": "",
        "code": "0019", "nature": "Sympt", "catalog": "COUGH", "standard": "Tuse",
        "domain": "respirator", "confidence": 0.88, "reason": "Cough expression."
    },
    {
        "patterns": ["transpir", "sudoare", "sudori"],
        "expr": "",
        "code": "0005", "nature": "Sympt", "catalog": "SWEATING", "standard": "Transpirații",
        "domain": "vegetativ", "confidence": 0.88, "reason": "Sweating expression."
    },
]


def _best_rule_suggestion(text):
    n = normalize_text(text)
    best = None
    for rule in RULE_SUGGESTIONS:
        for pat in rule["patterns"]:
            pn = normalize_text(pat)
            if pn and pn in n:
                cand = rule.copy()
                cand["matched_pattern"] = pat
                cand["expr"] = str(text).strip()
                if best is None or cand["confidence"] > best["confidence"]:
                    best = cand
    return best


def _best_fuzzy_suggestion(text, lexicon_runtime):
    if lexicon_runtime is None or lexicon_runtime.empty:
        return None

    n = normalize_text(text)
    if not n:
        return None

    best = None
    for _, r in lexicon_runtime.iterrows():
        expr = str(r.get("ExpresiePacient", "")).strip()
        expr_n = normalize_text(expr)
        if not expr_n:
            continue

        # Penalize very short expressions to avoid noisy suggestions.
        if len(expr_n) < 6:
            continue

        ratio = SequenceMatcher(None, n, expr_n).ratio()
        token_overlap = 0
        n_tokens = set(n.split())
        e_tokens = set(expr_n.split())
        if n_tokens and e_tokens:
            token_overlap = len(n_tokens & e_tokens) / max(1, len(e_tokens))

        score = max(ratio, token_overlap * 0.92)

        if score >= 0.74:
            if best is None or score > best["score"]:
                best = {
                    "score": score,
                    "expr": expr,
                    "code": clean_code(r.get("CodeElement", "")),
                    "nature": str(r.get("Nature Element", "")).strip(),
                    "catalog": str(r.get("CatalogName", "")).strip(),
                    "standard": str(r.get("ElementStandard", "")).strip(),
                    "domain": str(r.get("SemanticDomain", "")).strip(),
                    "reason": f"Fuzzy similarity to lexicon expression: {expr}",
                }

    return best


def autosuggest_for_unmatched(unmatched, lexicon_runtime):
    unmatched = ensure_columns(unmatched.copy(), SUGGESTION_COLUMNS)

    for i, row in unmatched.iterrows():
        # Preserve accepted/rejected/manual rows.
        validation = normalize_text(row.get("ValidationStatus", ""))
        if validation in {"accept", "accepted", "reject", "rejected", "manual", "edit", "edited"}:
            continue

        utterance = str(row.get("utterance", "") or row.get("text", "") or "").strip()
        candidate = str(row.get("CandidateExpression", "")).strip()
        source_text = candidate if candidate else utterance

        rule = _best_rule_suggestion(source_text)
        if rule:
            unmatched.at[i, "SuggestedExpression"] = candidate or rule["expr"]
            unmatched.at[i, "SuggestedElementStandard"] = rule["standard"]
            unmatched.at[i, "SuggestedCodeElement"] = rule["code"]
            unmatched.at[i, "SuggestedNatureElement"] = rule["nature"]
            unmatched.at[i, "SuggestedCatalogName"] = rule["catalog"]
            unmatched.at[i, "SemanticDomain"] = unmatched.at[i, "SemanticDomain"] or rule["domain"] if "SemanticDomain" in unmatched.columns else rule["domain"]
            unmatched.at[i, "SuggestionMethod"] = "rule"
            unmatched.at[i, "SuggestionConfidence"] = rule["confidence"]
            unmatched.at[i, "SuggestionReason"] = rule["reason"] + f" Matched pattern: {rule.get('matched_pattern', '')}"
            unmatched.at[i, "ValidationStatus"] = "suggested"
            continue

        fuzzy = _best_fuzzy_suggestion(source_text, lexicon_runtime)
        if fuzzy:
            unmatched.at[i, "SuggestedExpression"] = candidate or source_text
            unmatched.at[i, "SuggestedElementStandard"] = fuzzy["standard"]
            unmatched.at[i, "SuggestedCodeElement"] = fuzzy["code"]
            unmatched.at[i, "SuggestedNatureElement"] = fuzzy["nature"]
            unmatched.at[i, "SuggestedCatalogName"] = fuzzy["catalog"]
            unmatched.at[i, "SemanticDomain"] = unmatched.at[i, "SemanticDomain"] or fuzzy["domain"] if "SemanticDomain" in unmatched.columns else fuzzy["domain"]
            unmatched.at[i, "SuggestionMethod"] = "fuzzy"
            unmatched.at[i, "SuggestionConfidence"] = round(float(fuzzy["score"]), 3)
            unmatched.at[i, "SuggestionReason"] = fuzzy["reason"]
            unmatched.at[i, "ValidationStatus"] = "suggested"

    # Also fill Proposed* from accepted suggestions only.
    for i, row in unmatched.iterrows():
        validation = normalize_text(row.get("ValidationStatus", ""))
        if validation in {"accept", "accepted"}:
            if not str(row.get("CandidateExpression", "")).strip():
                unmatched.at[i, "CandidateExpression"] = row.get("SuggestedExpression", "")
            if not str(row.get("ProposedElementStandard", "")).strip():
                unmatched.at[i, "ProposedElementStandard"] = row.get("SuggestedElementStandard", "")
            if not str(row.get("ProposedCodeElement", "")).strip():
                unmatched.at[i, "ProposedCodeElement"] = row.get("SuggestedCodeElement", "")
            if not str(row.get("ProposedNatureElement", "")).strip():
                unmatched.at[i, "ProposedNatureElement"] = row.get("SuggestedNatureElement", "")
            if not str(row.get("ProposedCatalogName", "")).strip():
                unmatched.at[i, "ProposedCatalogName"] = row.get("SuggestedCatalogName", "")

    return unmatched


# --------------------------
# Lexicon integration
# --------------------------

def integrate_unmatched_into_lexicon(lex, unmatched):
    required = ["CandidateExpression", "ProposedCodeElement", "ProposedNatureElement", "ProposedCatalogName"]
    if unmatched is None or unmatched.empty:
        return lex, 0

    unmatched = ensure_columns(unmatched, required + SUGGESTION_COLUMNS)

    base_cols = [
        "ExprID", "ExpresiePacient", "ElementStandard", "CodeElement", "Nature Element",
        "CatalogName", "Intensitate", "Polaritate", "SemanticDomain", "Tags", "Observatii",
        "TermenAcademic", "TypeElement", "SourceSheets", "NormalizedExpression",
        "ReviewStatus", "ReviewerNote"
    ] + CONTEXT_COLUMNS
    lex = ensure_columns(lex, base_cols)

    existing = set(zip(
        lex["ExpresiePacient"].map(normalize_text),
        lex["CodeElement"].map(clean_code),
        lex["Nature Element"].astype(str),
        lex["CatalogName"].astype(str),
    ))

    new_rows = []
    for _, r in unmatched.fillna("").iterrows():
        validation = normalize_text(r.get("ValidationStatus", ""))

        # Only accepted rows are integrated automatically.
        if validation not in {"accept", "accepted"}:
            continue

        expr = str(r.get("CandidateExpression", "")).strip() or str(r.get("SuggestedExpression", "")).strip()
        code = clean_code(r.get("ProposedCodeElement", "") or r.get("SuggestedCodeElement", ""))
        nature = str(r.get("ProposedNatureElement", "") or r.get("SuggestedNatureElement", "")).strip()
        catalog = str(r.get("ProposedCatalogName", "") or r.get("SuggestedCatalogName", "")).strip()

        if not (expr and code and nature and catalog):
            continue

        key = (normalize_text(expr), code, nature, catalog)
        if key in existing:
            continue
        existing.add(key)

        elem = str(r.get("ProposedElementStandard", "") or r.get("SuggestedElementStandard", "")).strip()
        domain = str(r.get("SemanticDomain", "")).strip()

        row = {
            "ExprID": "",
            "ExpresiePacient": expr,
            "ElementStandard": elem,
            "CodeElement": code,
            "Nature Element": nature,
            "TypeElement": nature,
            "CatalogName": catalog,
            "Intensitate": "moderat",
            "Polaritate": "prezent",
            "Tags": "autosuggest_validated",
            "Observatii": "Imported from accepted AutoSuggestMapping row.",
            "TermenAcademic": elem,
            "SemanticDomain": domain,
            "SourceSheets": "Unmatched_AutoSuggest",
            "NormalizedExpression": normalize_text(expr),
            "ReviewStatus": "ImportedFromAcceptedSuggestion",
            "ReviewerNote": "",
        }

        for c in CONTEXT_COLUMNS:
            row[c] = str(r.get(c, "")).strip() if c in unmatched.columns else ""

        new_rows.append(row)

    if not new_rows:
        return lex, 0

    lex = pd.concat([lex, pd.DataFrame(new_rows)], ignore_index=True)
    if "ExprID" in lex.columns:
        lex["ExprID"] = range(1, len(lex) + 1)
    return lex, len(new_rows)


# --------------------------
# Annotation
# --------------------------

def annotate(lex, dialogues):
    required = ["ExpresiePacient", "CodeElement", "Nature Element", "CatalogName"]
    lex = ensure_columns(lex, required + CONTEXT_COLUMNS)
    lex["CodeElement"] = lex["CodeElement"].map(clean_code)
    lex["Nature Element"] = lex["Nature Element"].fillna("").astype(str).str.strip()
    lex["CatalogName"] = lex["CatalogName"].fillna("").astype(str).str.strip()
    lex["ExpresiePacient"] = lex["ExpresiePacient"].fillna("").astype(str).str.strip()

    runtime = lex[
        (lex["ExpresiePacient"] != "")
        & (lex["CodeElement"] != "")
        & (lex["Nature Element"] != "")
        & (lex["CatalogName"] != "")
    ].copy()

    runtime["expr_norm"] = runtime["ExpresiePacient"].map(normalize_text)
    runtime = runtime[runtime["expr_norm"].str.len() >= 3].copy()
    runtime["expr_len"] = runtime["expr_norm"].str.len()
    runtime = runtime.sort_values("expr_len", ascending=False).reset_index(drop=True)

    dialogues = ensure_columns(dialogues, ["dialog_id", "turn_id", "speaker", "text"])
    patient = dialogues[dialogues["speaker"].astype(str).str.lower() == "patient"].copy()
    patient["NormalizedText"] = patient["text"].map(normalize_text)

    rows = []
    ann_id = 1
    created = datetime.now().isoformat(timespec="seconds")

    for _, utt in patient.iterrows():
        u = normalize_text(utt.get("text", ""))
        if not u:
            continue

        matched = set()

        for _, lx in runtime.iterrows():
            e = lx["expr_norm"]
            words = e.split()
            method = ""
            conf = 0.0

            if e == u:
                method = "exact_full_utterance"; conf = 1.0
            elif len(words) >= 2 and e in u:
                method = "fragment_exact"; conf = 0.95
            elif len(words) == 1 and len(e) >= 6 and re.search(r"\b" + re.escape(e) + r"\b", u):
                method = "token_exact"; conf = 0.90
            else:
                continue

            key = (utt["dialog_id"], str(utt["turn_id"]), lx["CodeElement"], lx["Nature Element"])
            if key in matched:
                continue
            matched.add(key)

            detected_pol, neg_cue, pol_reason = detect_clinical_polarity(u, e, lx.get("Polaritate", "prezent"))
            detected_temp, temporal_cue, onset_type, duration_cat = detect_clinical_temporality(u)
            detected_sev, severity_cue, severity_score, severity_reason = detect_clinical_severity(u, lx.get("Intensitate", ""))

            row = {
                "annotation_id": f"ANN_{ann_id:06d}",
                "dialog_id": utt.get("dialog_id", ""),
                "turn_id": utt.get("turn_id", ""),
                "speaker": utt.get("speaker", ""),
                "utterance": utt.get("text", ""),
                "matched_expression": lx.get("ExpresiePacient", ""),
                "matched_expression_norm": e,
                "CodeElement": lx.get("CodeElement", ""),
                "NatureElement": lx.get("Nature Element", ""),
                "CatalogName": lx.get("CatalogName", ""),
                "ElementStandard": lx.get("ElementStandard", ""),
                "TermenAcademic": lx.get("TermenAcademic", ""),
                "SemanticDomain": lx.get("SemanticDomain", ""),
                "Intensitate": lx.get("Intensitate", ""),
                "Polaritate": lx.get("Polaritate", ""),
                "DetectedPolarity": detected_pol,
                "NegationCue": neg_cue,
                "PolarityReason": pol_reason,
                "DetectedTemporality": detected_temp,
                "TemporalCue": temporal_cue,
                "OnsetType": onset_type,
                "DurationCategory": duration_cat,
                "DetectedSeverity": detected_sev,
                "SeverityCue": severity_cue,
                "SeverityScore": severity_score,
                "SeverityReason": severity_reason,
            }

            for c in CONTEXT_COLUMNS:
                row[c] = lx.get(c, "")

            row.update({"MatchMethod": method, "Confidence": conf, "created_at": created})
            rows.append(row)
            ann_id += 1

    annotations = pd.DataFrame(rows)

    if annotations.empty:
        matched_keys = set()
    else:
        annotations["matched_len"] = annotations["matched_expression_norm"].str.len()
        annotations = annotations.sort_values(
            by=["dialog_id", "turn_id", "CodeElement", "Confidence", "matched_len"],
            ascending=[True, True, True, False, False],
        )
        annotations = annotations.drop_duplicates(
            subset=["dialog_id", "turn_id", "CodeElement", "NatureElement"],
            keep="first",
        )
        annotations = annotations.drop(columns=["matched_len"]).reset_index(drop=True)
        matched_keys = set(zip(annotations["dialog_id"].astype(str), annotations["turn_id"].astype(str)))

    unmatched = patient[
        ~patient.apply(lambda r: (str(r["dialog_id"]), str(r["turn_id"])) in matched_keys, axis=1)
    ].copy()

    for c in [
        "CandidateExpression", "ProposedElementStandard", "ProposedCodeElement",
        "ProposedNatureElement", "ProposedCatalogName", "SemanticDomain",
        "ReviewStatus", "ReviewerNote"
    ] + CONTEXT_COLUMNS + SUGGESTION_COLUMNS:
        unmatched[c] = ""

    return annotations, unmatched, runtime


# --------------------------
# Patient vectors / bridge / synthesis
# --------------------------

def _join_unique(values):
    seen = []
    for v in values:
        s = "" if v is None else str(v).strip()
        if not s or s.lower() == "nan":
            continue
        if s not in seen:
            seen.append(s)
    return ";".join(seen)


def _subset_codes(df, nature, polarity):
    if df.empty:
        return ""
    sub = df[(df["NatureElement"].astype(str) == nature) & (df["DetectedPolarity"].astype(str) == polarity)]
    return _join_unique(sub["CodeElement"].astype(str).tolist())


def _subset_labels(df, nature, polarity):
    if df.empty:
        return ""
    sub = df[(df["NatureElement"].astype(str) == nature) & (df["DetectedPolarity"].astype(str) == polarity)]
    pairs = []
    for _, r in sub.iterrows():
        code = str(r.get("CodeElement", "")).strip()
        name = str(r.get("CatalogName", "")).strip()
        if code and name:
            pairs.append(f"{code}:{name}")
    return _join_unique(pairs)


def build_patient_vectors(annotations):
    if annotations is None or annotations.empty:
        return pd.DataFrame()

    df = annotations.copy()
    for c in ["dialog_id", "CodeElement", "NatureElement", "CatalogName", "DetectedPolarity", "SeverityScore", "utterance", "matched_expression"]:
        if c not in df.columns:
            df[c] = ""

    rows = []
    for dialog_id, g in df.groupby("dialog_id", dropna=False):
        present = g[g["DetectedPolarity"].astype(str) == "prezent"].copy()
        absent = g[g["DetectedPolarity"].astype(str) == "absent"].copy()
        resolved = g[g["DetectedPolarity"].astype(str) == "rezolvat"].copy()

        severity_scores = []
        for v in present.get("SeverityScore", []):
            try:
                severity_scores.append(int(float(v)))
            except Exception:
                pass
        sev_max = max(severity_scores) if severity_scores else ""

        fragments = []
        for _, r in g.iterrows():
            code = str(r.get("CodeElement", "")).strip()
            pol = str(r.get("DetectedPolarity", "")).strip()
            frag = str(r.get("matched_expression", "")).strip()
            utt = str(r.get("utterance", "")).strip()
            if code and frag:
                fragments.append(f"{code}|{pol}|{frag}|{utt}")

        rows.append({
            "dialog_id": dialog_id,
            "PresentSymptCodes": _subset_codes(present, "Sympt", "prezent"),
            "PresentSigneCodes": _subset_codes(present, "Signe", "prezent"),
            "PresentRiskFCodes": _subset_codes(present, "RiskF", "prezent"),
            "AbsentCodes": _join_unique(absent["CodeElement"].astype(str).tolist()),
            "ResolvedCodes": _join_unique(resolved["CodeElement"].astype(str).tolist()),
            "PresentSymptLabels": _subset_labels(present, "Sympt", "prezent"),
            "PresentSigneLabels": _subset_labels(present, "Signe", "prezent"),
            "PresentRiskFLabels": _subset_labels(present, "RiskF", "prezent"),
            "AbsentLabels": _join_unique([f"{r.get('CodeElement','')}:{r.get('CatalogName','')}" for _, r in absent.iterrows()]),
            "ResolvedLabels": _join_unique([f"{r.get('CodeElement','')}:{r.get('CatalogName','')}" for _, r in resolved.iterrows()]),
            "ContextSummary": "",
            "TemporalSummary": "",
            "SeverityMax": sev_max,
            "SeveritySummary": "",
            "SourceFragments": _join_unique(fragments),
            "N_Present": len(present),
            "N_Absent": len(absent),
            "N_Resolved": len(resolved),
            "VectorStatus": "ready_for_mediag" if len(present) > 0 else "no_present_elements",
        })

    return pd.DataFrame(rows).sort_values("dialog_id").reset_index(drop=True)


def _codes_to_list(value):
    if value is None:
        return []
    s = str(value).strip()
    if not s or s.lower() == "nan":
        return []
    out = []
    for x in re.split(r"[;,]+", s):
        x = x.strip()
        if x and x not in out:
            out.append(x)
    return out


def build_mediag_input(patient_vectors):
    if patient_vectors is None or patient_vectors.empty:
        return pd.DataFrame()

    rows = []
    for _, r in patient_vectors.iterrows():
        sympt = _codes_to_list(r.get("PresentSymptCodes", ""))
        signe = _codes_to_list(r.get("PresentSigneCodes", ""))
        riskf = _codes_to_list(r.get("PresentRiskFCodes", ""))
        absent = _codes_to_list(r.get("AbsentCodes", ""))
        resolved = _codes_to_list(r.get("ResolvedCodes", ""))

        payload = {
            "dialog_id": str(r.get("dialog_id", "")),
            "sympt": sympt,
            "signe": signe,
            "riskf": riskf,
            "absent": absent,
            "resolved": resolved,
            "context_summary": str(r.get("ContextSummary", "")),
            "temporal_summary": str(r.get("TemporalSummary", "")),
            "severity_max": str(r.get("SeverityMax", "")),
            "severity_summary": str(r.get("SeveritySummary", "")),
            "source_fragments": str(r.get("SourceFragments", "")),
        }

        rows.append({
            "dialog_id": str(r.get("dialog_id", "")),
            "SymptCodes": ",".join(sympt),
            "SigneCodes": ",".join(signe),
            "RiskFCodes": ",".join(riskf),
            "AbsentCodes": ",".join(absent),
            "ResolvedCodes": ",".join(resolved),
            "N_Sympt": len(sympt),
            "N_Signe": len(signe),
            "N_RiskF": len(riskf),
            "N_Absent": len(absent),
            "N_Resolved": len(resolved),
            "ContextSummary": str(r.get("ContextSummary", "")),
            "TemporalSummary": str(r.get("TemporalSummary", "")),
            "SeverityMax": str(r.get("SeverityMax", "")),
            "SeveritySummary": str(r.get("SeveritySummary", "")),
            "SourceFragments": str(r.get("SourceFragments", "")),
            "MedDiagPayloadJSON": json.dumps(payload, ensure_ascii=False),
            "BridgeStatus": "ready_for_mediag" if len(sympt) + len(signe) + len(riskf) > 0 else "no_active_codes",
        })

    return pd.DataFrame(rows)


def build_clinical_synthesis(patient_vectors, mediag_input):
    if patient_vectors is None or patient_vectors.empty:
        return pd.DataFrame()

    rows = []
    for _, r in patient_vectors.iterrows():
        labels = []
        for col, title in [("PresentSymptLabels", "Symptômes présents"), ("PresentSigneLabels", "Signes présents"), ("PresentRiskFLabels", "Facteurs de risque"), ("AbsentLabels", "Éléments absents"), ("ResolvedLabels", "Éléments résolus")]:
            val = str(r.get(col, "")).strip()
            if val:
                labels.append(f"{title}: {val}.")
        synthesis = " ".join(labels) if labels else "Aucun élément clinique actif suffisamment codé n'a été extrait."

        rows.append({
            "dialog_id": str(r.get("dialog_id", "")),
            "ClinicalSynthesis": synthesis,
            "ClinicalOrientation": "orientation clinique à préciser",
            "TriageHint": "priorité à évaluer selon contexte clinique complet",
            "SymptCodes": "",
            "SigneCodes": "",
            "RiskFCodes": "",
            "AbsentCodes": str(r.get("AbsentCodes", "")),
            "ResolvedCodes": str(r.get("ResolvedCodes", "")),
            "ContextSummary": str(r.get("ContextSummary", "")),
            "TemporalSummary": str(r.get("TemporalSummary", "")),
            "SeverityMax": str(r.get("SeverityMax", "")),
            "SeveritySummary": str(r.get("SeveritySummary", "")),
            "SourceFragments": str(r.get("SourceFragments", "")),
            "MedDiagPayloadJSON": "",
            "SynthesisStatus": "ready_for_clinical_review" if str(r.get("VectorStatus", "")) == "ready_for_mediag" else "insufficient_extracted_elements",
        })

    out = pd.DataFrame(rows)
    if mediag_input is not None and not mediag_input.empty:
        bridge_cols = ["dialog_id", "SymptCodes", "SigneCodes", "RiskFCodes", "AbsentCodes", "ResolvedCodes", "MedDiagPayloadJSON"]
        out = out.drop(columns=[c for c in bridge_cols if c in out.columns and c != "dialog_id"], errors="ignore")
        out = out.merge(mediag_input[bridge_cols], on="dialog_id", how="left")
    return out


# --------------------------
# AutoSuggestions sheet for ALL patient utterances
# --------------------------

def _annotation_keys_by_dialog_turn(annotations):
    out = {}
    if annotations is None or annotations.empty:
        return out

    for _, r in annotations.iterrows():
        key = (str(r.get("dialog_id", "")), str(r.get("turn_id", "")))
        item = (
            str(r.get("CodeElement", "")),
            str(r.get("CatalogName", "")),
            str(r.get("matched_expression", "")),
        )
        out.setdefault(key, []).append(item)
    return out


def _split_candidate_phrases(text):
    """
    Conservative phrase splitting for auto-suggestion.
    Keeps the full utterance and also subphrases split by punctuation / connectors.
    """
    raw = str(text or "").strip()
    if not raw:
        return []

    candidates = [raw]

    # Split on punctuation and common connectors but keep meaningful pieces.
    chunks = re.split(r"[.;!?]+|\bsi\b|\bși\b|\bdar\b|\biar\b|\bplus\b", raw, flags=re.IGNORECASE)
    for ch in chunks:
        ch = ch.strip(" ,;:.!?")
        if len(normalize_text(ch)) >= 5:
            candidates.append(ch)

    # Remove duplicates while preserving order.
    seen = []
    for c in candidates:
        cn = normalize_text(c)
        if cn and cn not in seen:
            seen.append(cn)

    final = []
    for cn in seen:
        # reconstruct normalized phrase as acceptable SuggestedExpression source
        final.append(cn)
    return final


def build_autosuggestions_all(dialogues, annotations, lexicon_runtime):
    """
    Generates suggestions for every patient utterance, even if the utterance already has
    one or more annotations. This solves the problem where a long patient answer is
    partially matched and therefore never appears in Unmatched.
    """
    columns = [
        "dialog_id",
        "turn_id",
        "speaker",
        "utterance",
        "CandidateExpression",
        "AlreadyMatchedCodes",
        "SuggestedExpression",
        "SuggestedElementStandard",
        "SuggestedCodeElement",
        "SuggestedNatureElement",
        "SuggestedCatalogName",
        "SuggestionMethod",
        "SuggestionConfidence",
        "SuggestionReason",
        "ValidationStatus",
        "ProposedElementStandard",
        "ProposedCodeElement",
        "ProposedNatureElement",
        "ProposedCatalogName",
        "SemanticDomain",
    ]

    if dialogues is None or dialogues.empty:
        return pd.DataFrame(columns=columns)

    dialogues = ensure_columns(dialogues.copy(), ["dialog_id", "turn_id", "speaker", "text"])
    patient = dialogues[dialogues["speaker"].astype(str).str.lower() == "patient"].copy()

    ann_map = _annotation_keys_by_dialog_turn(annotations)

    rows = []
    seen_suggestions = set()

    for _, utt in patient.iterrows():
        dialog_id = str(utt.get("dialog_id", ""))
        turn_id = str(utt.get("turn_id", ""))
        utterance = str(utt.get("text", "")).strip()
        key = (dialog_id, turn_id)

        already = ann_map.get(key, [])
        already_codes = ";".join([f"{c}:{name}" for c, name, expr in already if c])

        candidate_phrases = _split_candidate_phrases(utterance)

        for cand in candidate_phrases:
            # Skip candidates that are already exactly covered by matched expressions.
            cand_norm = normalize_text(cand)
            matched_exprs = [normalize_text(expr) for _, _, expr in already]
            if any(m and m in cand_norm for m in matched_exprs):
                # Do not skip the full utterance if it contains other clinical clues.
                if cand_norm == normalize_text(utterance):
                    pass
                else:
                    continue

            rule = _best_rule_suggestion(cand)
            suggestion = None

            if rule:
                suggestion = {
                    "SuggestedExpression": cand,
                    "SuggestedElementStandard": rule["standard"],
                    "SuggestedCodeElement": rule["code"],
                    "SuggestedNatureElement": rule["nature"],
                    "SuggestedCatalogName": rule["catalog"],
                    "SuggestionMethod": "rule_all_utterances",
                    "SuggestionConfidence": rule["confidence"],
                    "SuggestionReason": rule["reason"] + f" Matched pattern: {rule.get('matched_pattern', '')}",
                    "SemanticDomain": rule["domain"],
                }
            else:
                fuzzy = _best_fuzzy_suggestion(cand, lexicon_runtime)
                if fuzzy:
                    suggestion = {
                        "SuggestedExpression": cand,
                        "SuggestedElementStandard": fuzzy["standard"],
                        "SuggestedCodeElement": fuzzy["code"],
                        "SuggestedNatureElement": fuzzy["nature"],
                        "SuggestedCatalogName": fuzzy["catalog"],
                        "SuggestionMethod": "fuzzy_all_utterances",
                        "SuggestionConfidence": round(float(fuzzy["score"]), 3),
                        "SuggestionReason": fuzzy["reason"],
                        "SemanticDomain": fuzzy["domain"],
                    }

            if not suggestion:
                continue

            # Avoid suggesting codes already matched in that same turn, unless the suggested expression is a distinct phrase.
            suggested_code = str(suggestion["SuggestedCodeElement"])
            if suggested_code in [c for c, _, _ in already]:
                # keep if phrase is not already matched and confidence high
                try:
                    conf = float(suggestion["SuggestionConfidence"])
                except Exception:
                    conf = 0
                if conf < 0.94:
                    continue

            dedup = (dialog_id, turn_id, normalize_text(suggestion["SuggestedExpression"]), suggested_code)
            if dedup in seen_suggestions:
                continue
            seen_suggestions.add(dedup)

            rows.append({
                "dialog_id": dialog_id,
                "turn_id": turn_id,
                "speaker": "patient",
                "utterance": utterance,
                "CandidateExpression": suggestion["SuggestedExpression"],
                "AlreadyMatchedCodes": already_codes,
                "SuggestedExpression": suggestion["SuggestedExpression"],
                "SuggestedElementStandard": suggestion["SuggestedElementStandard"],
                "SuggestedCodeElement": suggestion["SuggestedCodeElement"],
                "SuggestedNatureElement": suggestion["SuggestedNatureElement"],
                "SuggestedCatalogName": suggestion["SuggestedCatalogName"],
                "SuggestionMethod": suggestion["SuggestionMethod"],
                "SuggestionConfidence": suggestion["SuggestionConfidence"],
                "SuggestionReason": suggestion["SuggestionReason"],
                "ValidationStatus": "suggested",
                "ProposedElementStandard": "",
                "ProposedCodeElement": "",
                "ProposedNatureElement": "",
                "ProposedCatalogName": "",
                "SemanticDomain": suggestion["SemanticDomain"],
            })

    if not rows:
        return pd.DataFrame(columns=columns)

    return pd.DataFrame(rows, columns=columns)


def integrate_accepted_autosuggestions_into_lexicon(lex, autosuggestions):
    """
    Imports accepted suggestions from AutoSuggestions sheet into Lexicon.
    User sets ValidationStatus = accept.
    """
    if autosuggestions is None or autosuggestions.empty:
        return lex, 0

    autosuggestions = ensure_columns(autosuggestions.copy(), [
        "CandidateExpression",
        "SuggestedExpression",
        "SuggestedElementStandard",
        "SuggestedCodeElement",
        "SuggestedNatureElement",
        "SuggestedCatalogName",
        "ValidationStatus",
        "ProposedElementStandard",
        "ProposedCodeElement",
        "ProposedNatureElement",
        "ProposedCatalogName",
        "SemanticDomain",
    ])

    # Convert accepted suggestions into Unmatched-like rows and reuse integration.
    rows = []
    for _, r in autosuggestions.iterrows():
        validation = normalize_text(r.get("ValidationStatus", ""))
        if validation not in {"accept", "accepted"}:
            continue

        expr = str(r.get("CandidateExpression", "") or r.get("SuggestedExpression", "")).strip()
        elem = str(r.get("ProposedElementStandard", "") or r.get("SuggestedElementStandard", "")).strip()
        code = str(r.get("ProposedCodeElement", "") or r.get("SuggestedCodeElement", "")).strip()
        nature = str(r.get("ProposedNatureElement", "") or r.get("SuggestedNatureElement", "")).strip()
        catalog = str(r.get("ProposedCatalogName", "") or r.get("SuggestedCatalogName", "")).strip()
        domain = str(r.get("SemanticDomain", "")).strip()

        if not (expr and code and nature and catalog):
            continue

        rows.append({
            "CandidateExpression": expr,
            "ProposedElementStandard": elem,
            "ProposedCodeElement": code,
            "ProposedNatureElement": nature,
            "ProposedCatalogName": catalog,
            "SemanticDomain": domain,
            "ValidationStatus": "accept",
        })

    if not rows:
        return lex, 0

    pseudo_unmatched = pd.DataFrame(rows)
    return integrate_unmatched_into_lexicon(lex, pseudo_unmatched)


# --------------------------
# Runtime / match audit
# --------------------------

def _audit_required_value(value):
    s = "" if value is None else str(value).strip()
    if not s or s.lower() in {"nan", "none", "null"}:
        return ""
    return s


def build_lexicon_runtime_audit(lex, dialogues):
    """
    Shows why each lexicon row is included or excluded from runtime.
    Also checks whether the normalized expression is textually present in any patient utterance.
    """
    if lex is None or lex.empty:
        return pd.DataFrame()

    df = lex.copy()
    df = ensure_columns(df, [
        "ExprID",
        "ExpresiePacient",
        "CodeElement",
        "Nature Element",
        "TypeElement",
        "CatalogName",
        "ElementStandard",
        "SemanticDomain",
        "ReviewStatus",
        "SourceSheets",
    ])

    patient = pd.DataFrame()
    if dialogues is not None and not dialogues.empty:
        d = ensure_columns(dialogues.copy(), ["dialog_id", "turn_id", "speaker", "text"])
        patient = d[d["speaker"].astype(str).str.lower() == "patient"].copy()
        patient["utterance_norm"] = patient["text"].map(normalize_text)

    rows = []
    for idx, r in df.iterrows():
        expr = _audit_required_value(r.get("ExpresiePacient", ""))
        code = clean_code(r.get("CodeElement", ""))
        nature = _audit_required_value(r.get("Nature Element", ""))
        catalog = _audit_required_value(r.get("CatalogName", ""))
        expr_norm = normalize_text(expr)

        missing = []
        if not expr:
            missing.append("ExpresiePacient")
        if not code:
            missing.append("CodeElement")
        if not nature:
            missing.append("Nature Element")
        if not catalog:
            missing.append("CatalogName")
        if expr and len(expr_norm) < 3:
            missing.append("NormalizedExpression too short")

        included = len(missing) == 0

        matched_dialogs = []
        matched_turns = []
        matched_utterances = []

        if included and expr_norm and not patient.empty:
            for _, u in patient.iterrows():
                utterance_norm = str(u.get("utterance_norm", ""))
                if expr_norm in utterance_norm:
                    did = str(u.get("dialog_id", ""))
                    tid = str(u.get("turn_id", ""))
                    utt = str(u.get("text", ""))
                    matched_dialogs.append(did)
                    matched_turns.append(f"{did}:{tid}")
                    matched_utterances.append(utt)

        rows.append({
            "ExcelRowApprox": int(idx) + 2,
            "ExprID": r.get("ExprID", ""),
            "ExpresiePacient": expr,
            "NormalizedExpression": expr_norm,
            "CodeElement": code,
            "NatureElement": nature,
            "TypeElement": r.get("TypeElement", ""),
            "CatalogName": catalog,
            "ElementStandard": r.get("ElementStandard", ""),
            "SemanticDomain": r.get("SemanticDomain", ""),
            "IncludedInRuntime": "YES" if included else "NO",
            "ExclusionReason": "" if included else "; ".join(missing),
            "MatchedAnyPatientUtterance": "YES" if matched_dialogs else "NO",
            "MatchedDialogIDs": ";".join(sorted(set(matched_dialogs))),
            "MatchedTurns": ";".join(matched_turns[:10]),
            "MatchedUtteranceExamples": " || ".join(matched_utterances[:3]),
            "ReviewStatus": r.get("ReviewStatus", ""),
            "SourceSheets": r.get("SourceSheets", ""),
        })

    return pd.DataFrame(rows)


def build_lexicon_dialog_match_audit(lex_runtime, dialogues):
    """
    Lists actual lexicon-expression matches against patient utterances.
    Useful to debug a selected dialog such as LONG_NEURO_PARKINSONISM_0001.
    """
    cols = [
        "dialog_id", "turn_id", "utterance", "ExprID", "ExpresiePacient",
        "CodeElement", "NatureElement", "CatalogName", "ElementStandard", "MatchType"
    ]

    if lex_runtime is None or lex_runtime.empty or dialogues is None or dialogues.empty:
        return pd.DataFrame(columns=cols)

    d = ensure_columns(dialogues.copy(), ["dialog_id", "turn_id", "speaker", "text"])
    patient = d[d["speaker"].astype(str).str.lower() == "patient"].copy()
    patient["utterance_norm"] = patient["text"].map(normalize_text)

    rows = []
    for _, u in patient.iterrows():
        utterance_norm = str(u.get("utterance_norm", ""))
        if not utterance_norm:
            continue

        for _, lx in lex_runtime.iterrows():
            expr_norm = str(lx.get("expr_norm", ""))
            if expr_norm and expr_norm in utterance_norm:
                rows.append({
                    "dialog_id": u.get("dialog_id", ""),
                    "turn_id": u.get("turn_id", ""),
                    "utterance": u.get("text", ""),
                    "ExprID": lx.get("ExprID", ""),
                    "ExpresiePacient": lx.get("ExpresiePacient", ""),
                    "CodeElement": lx.get("CodeElement", ""),
                    "NatureElement": lx.get("Nature Element", ""),
                    "CatalogName": lx.get("CatalogName", ""),
                    "ElementStandard": lx.get("ElementStandard", ""),
                    "MatchType": "fragment_exact",
                })

    return pd.DataFrame(rows, columns=cols)


# --------------------------
# Main
# --------------------------

def main():
    if not WORKBOOK.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK}")

    sheets = pd.read_excel(WORKBOOK, sheet_name=None, dtype=str)
    sheets = {k: v.fillna("") for k, v in sheets.items()}

    lex = sheets.get("Lexicon", pd.DataFrame()).fillna("")
    dialogues = sheets.get("Dialogues", pd.DataFrame()).fillna("")
    old_unmatched = sheets.get("Unmatched", pd.DataFrame()).fillna("")
    old_autosuggestions = sheets.get("AutoSuggestions", pd.DataFrame()).fillna("")

    lex, imported_unmatched = integrate_unmatched_into_lexicon(lex, old_unmatched)
    lex, imported_autosuggest = integrate_accepted_autosuggestions_into_lexicon(lex, old_autosuggestions)
    imported = imported_unmatched + imported_autosuggest

    annotations, unmatched, runtime = annotate(lex, dialogues)
    unmatched = autosuggest_for_unmatched(unmatched, runtime)
    autosuggestions = build_autosuggestions_all(dialogues, annotations, runtime)
    lexicon_runtime_audit = build_lexicon_runtime_audit(lex, dialogues)
    lexicon_dialog_match_audit = build_lexicon_dialog_match_audit(runtime, dialogues)

    patient_vectors = build_patient_vectors(annotations)
    mediag_input = build_mediag_input(patient_vectors)
    clinical_synthesis = build_clinical_synthesis(patient_vectors, mediag_input)

    suggestions = int((unmatched.get("ValidationStatus", pd.Series(dtype=str)).astype(str) == "suggested").sum()) if not unmatched.empty else 0

    summary = pd.DataFrame([
        {"Metric": "run_at", "Value": datetime.now().isoformat(timespec="seconds")},
        {"Metric": "imported_from_accepted_suggestions", "Value": imported},
        {"Metric": "lexicon_rows", "Value": len(lex)},
        {"Metric": "runtime_coded_lexicon_rows", "Value": len(runtime)},
        {"Metric": "dialogue_turns", "Value": len(dialogues)},
        {"Metric": "annotations", "Value": len(annotations)},
        {"Metric": "unmatched_patient_utterances", "Value": len(unmatched)},
        {"Metric": "unmatched_autosuggestions", "Value": suggestions},
        {"Metric": "all_utterance_autosuggestions", "Value": len(autosuggestions)},
        {"Metric": "patient_vectors", "Value": len(patient_vectors)},
        {"Metric": "mediag_bridge_rows", "Value": len(mediag_input)},
        {"Metric": "clinical_synthesis_rows", "Value": len(clinical_synthesis)},
        {"Metric": "lexicon_runtime_audit_rows", "Value": len(lexicon_runtime_audit)},
        {"Metric": "lexicon_dialog_match_audit_rows", "Value": len(lexicon_dialog_match_audit)},
    ])

    with pd.ExcelWriter(WORKBOOK, engine="openpyxl", mode="w") as writer:
        lex.to_excel(writer, index=False, sheet_name="Lexicon")
        dialogues.to_excel(writer, index=False, sheet_name="Dialogues")
        annotations.to_excel(writer, index=False, sheet_name="Annotations")
        patient_vectors.to_excel(writer, index=False, sheet_name="PatientVectors")
        mediag_input.to_excel(writer, index=False, sheet_name="MedDiagInput")
        clinical_synthesis.to_excel(writer, index=False, sheet_name="ClinicalSynthesis")
        autosuggestions.to_excel(writer, index=False, sheet_name="AutoSuggestions")
        lexicon_runtime_audit.to_excel(writer, index=False, sheet_name="LexiconRuntimeAudit")
        lexicon_dialog_match_audit.to_excel(writer, index=False, sheet_name="LexiconDialogMatchAudit")
        unmatched.to_excel(writer, index=False, sheet_name="Unmatched")
        summary.to_excel(writer, index=False, sheet_name="Summary")

        if "LoaderLog" in sheets:
            sheets["LoaderLog"].to_excel(writer, index=False, sheet_name="LoaderLog")
        if "MappingAudit_V13" in sheets:
            sheets["MappingAudit_V13"].to_excel(writer, index=False, sheet_name="MappingAudit_V13")

    format_workbook(WORKBOOK)

    print("READY")
    print(f"Workbook: {WORKBOOK}")
    print(f"Imported accepted suggestions: {imported}")
    print(f"Runtime coded lexicon rows: {len(runtime)}")
    print(f"Annotations: {len(annotations)}")
    print(f"Unmatched: {len(unmatched)}")
    print(f"Unmatched auto-suggestions: {suggestions}")
    print(f"All-utterance auto-suggestions: {len(autosuggestions)}")
    print(f"Patient vectors: {len(patient_vectors)}")
    print(f"MedDiag rows: {len(mediag_input)}")
    print(f"LexiconRuntimeAudit rows: {len(lexicon_runtime_audit)}")
    print(f"LexiconDialogMatchAudit rows: {len(lexicon_dialog_match_audit)}")


if __name__ == "__main__":
    main()
