# DOR_NlpRO_streamlit_app_v12_runtime_audit.py
# Filename: DOR_NlpRO_streamlit_app_v12_runtime_audit.py
#
# NlpRO - Romanian Clinical NLP Streamlit Interface
#
# Expected files in the same folder:
#   - ClinicalPipeline_RO_SINGLE_v14_RUNTIME_AUDIT.xlsx
#   - DOR_clinical_pipeline_singlefile_v14_runtime_audit.py
#
# Install:
#   pip install streamlit pandas openpyxl
#
# Run:
#   streamlit run DOR_NlpRO_streamlit_app_v01.py

from pathlib import Path
from datetime import datetime
import io
import json
import re
import subprocess
import sys

import pandas as pd
import streamlit as st
from ANAM_titus_tab import render_titus_tab


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_WORKBOOK = BASE_DIR / "ClinicalPipeline_RO_SINGLE_v14_RUNTIME_AUDIT.xlsx"
DEFAULT_PIPELINE_SCRIPT = BASE_DIR / "DOR_clinical_pipeline_singlefile_v17_SAFE_BACKUP.py"

st.set_page_config(page_title="NlpRO", page_icon="🩺", layout="wide")


def normalize_speaker(raw):
    s = str(raw).strip().lower()
    aliases = {
        "doctor": "doctor",
        "dr": "doctor",
        "medic": "doctor",
        "medicul": "doctor",
        "patient": "patient",
        "pacient": "patient",
        "pacienta": "patient",
        "pacientă": "patient",
        "bolnav": "patient",
        "bolnavul": "patient",
    }
    return aliases.get(s, s)


def parse_transcript_text(text):
    text = text.replace("\ufeff", "")
    lines = text.splitlines()

    dialog_id = ""
    case_title = ""
    turns = []

    current_speaker = None
    current_text_parts = []

    def flush_current():
        nonlocal current_speaker, current_text_parts, turns
        if current_speaker and current_text_parts:
            clean_text = " ".join(x.strip() for x in current_text_parts if x.strip()).strip()
            if clean_text:
                turns.append((current_speaker, clean_text))
        current_speaker = None
        current_text_parts = []

    for raw_line in lines:
        line = raw_line.strip()

        if not line:
            continue

        m_case = re.match(r"^(CASE|CAZ|TITLE|TITLU)\s*:\s*(.+)$", line, flags=re.IGNORECASE)
        if m_case:
            case_title = m_case.group(2).strip()
            continue

        m_id = re.match(r"^(DIALOG_ID|ID|DIALOGUE_ID)\s*:\s*(.+)$", line, flags=re.IGNORECASE)
        if m_id:
            dialog_id = m_id.group(2).strip()
            continue

        m_turn = re.match(
            r"^(Doctor|Dr|Medic|Medicul|Patient|Pacient|Pacienta|Pacientă|Bolnav|Bolnavul)\s*:\s*(.*)$",
            line,
            flags=re.IGNORECASE,
        )
        if m_turn:
            flush_current()
            current_speaker = normalize_speaker(m_turn.group(1))
            first_text = m_turn.group(2).strip()
            if first_text:
                current_text_parts.append(first_text)
            continue

        if current_speaker:
            current_text_parts.append(line)

    flush_current()

    if not dialog_id:
        dialog_id = f"DLG_IMPORTED_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

    rows = []
    for i, (speaker, utterance) in enumerate(turns, start=1):
        rows.append({
            "dialog_id": dialog_id,
            "turn_id": i,
            "speaker": speaker,
            "text": utterance,
            "theme": case_title,
            "specialty": "",
            "case_type": "streamlit_loaded_transcript",
            "language": "ro",
            "NormalizedText": "",
        })

    return dialog_id, case_title, pd.DataFrame(rows)


@st.cache_data(show_spinner=False)
def load_workbook(path_str):
    path = Path(path_str)
    if not path.exists():
        return {}

    try:
        sheets = pd.read_excel(path, sheet_name=None, dtype=str)
    except Exception as exc:
        st.error(f"Nu pot citi workbook-ul: {exc}")
        return {}

    return {name: df.fillna("") for name, df in sheets.items()}


def _wb_get(workbook_path_str):
    """Returneaza sheets din session_state daca disponibil, altfel din cache."""
    key = f"wb_sheets_{workbook_path_str}"
    if key not in st.session_state:
        st.session_state[key] = load_workbook(workbook_path_str).copy()
    return st.session_state[key]


def _wb_update_sheet(workbook_path_str, sheet_name, df):
    """Actualizeaza un singur sheet in session_state — fara reread de pe disc."""
    key = f"wb_sheets_{workbook_path_str}"
    if key not in st.session_state:
        st.session_state[key] = load_workbook(workbook_path_str).copy()
    st.session_state[key][sheet_name] = df.fillna("")


def _wb_invalidate(workbook_path_str):
    """Invalideaza session_state si cache — forteaza reread de pe disc."""
    key = f"wb_sheets_{workbook_path_str}"
    if key in st.session_state:
        del st.session_state[key]
    _wb_invalidate(str(workbook_path))


def save_workbook(path, sheets):
    preferred = [
        "Lexicon",
        "Dialogues",
        "Annotations",
        "PatientVectors",
        "MedDiagInput",
        "ClinicalSynthesis",
        "Unmatched",
        "Summary",
        "LoaderLog",
    ]

    with pd.ExcelWriter(path, engine="openpyxl", mode="w") as writer:
        written = set()

        for name in preferred:
            if name in sheets:
                sheets[name].to_excel(writer, index=False, sheet_name=name)
                written.add(name)

        for name, df in sheets.items():
            if name not in written:
                df.to_excel(writer, index=False, sheet_name=name)


def append_dialogue_to_workbook(workbook_path, new_dialogue, dialog_id, case_title, replace_existing=True):
    sheets = load_workbook(str(workbook_path)).copy()

    if "Dialogues" not in sheets:
        raise ValueError("Workbook-ul nu conține foaia Dialogues.")

    dialogues = sheets["Dialogues"].copy()

    required_columns = [
        "dialog_id",
        "turn_id",
        "speaker",
        "text",
        "theme",
        "specialty",
        "case_type",
        "language",
        "NormalizedText",
    ]

    for col in required_columns:
        if col not in dialogues.columns:
            dialogues[col] = ""

    new_dialogue = new_dialogue[required_columns]

    if replace_existing:
        before = len(dialogues)
        dialogues = dialogues[dialogues["dialog_id"].astype(str) != dialog_id].copy()
        removed = before - len(dialogues)
    else:
        removed = 0

    dialogues = pd.concat([dialogues, new_dialogue], ignore_index=True)
    sheets["Dialogues"] = dialogues

    log_row = pd.DataFrame([{
        "loaded_at": datetime.now().isoformat(timespec="seconds"),
        "dialog_id": dialog_id,
        "case_title": case_title,
        "turns_loaded": len(new_dialogue),
        "previous_rows_removed_same_dialog_id": removed,
        "source": "NlpRO Streamlit",
    }])

    if "LoaderLog" in sheets:
        sheets["LoaderLog"] = pd.concat([sheets["LoaderLog"], log_row], ignore_index=True)
    else:
        sheets["LoaderLog"] = log_row

    save_workbook(workbook_path, sheets)
    _wb_invalidate(str(workbook_path))

    return removed, len(new_dialogue)


def run_pipeline(script_path):
    if not script_path.exists():
        raise FileNotFoundError(f"Nu găsesc scriptul pipeline: {script_path}")

    result = subprocess.run(
        [sys.executable, str(script_path)],
        cwd=str(script_path.parent),
        capture_output=True,
        text=True,
    )

    # Pipeline-ul modifica workbook-ul pe disc — invalidam session_state
    # dar NU load_workbook cache (evitam reread la rerun imediat)
    # Invalidarea se face explicit de catre apelant daca e necesar
    return result.returncode, result.stdout, result.stderr


def dataframe_download_button(df, filename, label):
    buffer = io.BytesIO()
    df.to_excel(buffer, index=False, engine="openpyxl")
    st.download_button(
        label=label,
        data=buffer.getvalue(),
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def severity_label_from_score(score):
    try:
        s = int(float(score))
    except Exception:
        return ""
    if s >= 3:
        return "puternic"
    if s == 2:
        return "moderat"
    if s == 1:
        return "slab"
    return ""


def build_mediag_display_table(active_annotations):
    """
    Build a clinically readable table for the selected dialogue:
      CodeElement + NatureElement + CatalogName + polarity + estimated intensity/severity.
    """
    if active_annotations is None or active_annotations.empty:
        return pd.DataFrame(columns=[
            "Type",
            "Code",
            "Element",
            "Polarity",
            "EstimatedIntensity",
            "TemporalContext",
            "ClinicalContext",
            "MatchedExpression",
        ])

    df = active_annotations.copy()

    for col in [
        "NatureElement", "CodeElement", "CatalogName", "DetectedPolarity",
        "DetectedSeverity", "SeverityScore", "Intensitate",
        "DetectedTemporality", "TemporalContext", "TriggerContext",
        "BodyRegion", "ClinicalIntent", "matched_expression"
    ]:
        if col not in df.columns:
            df[col] = ""

    rows = []
    seen = set()

    for _, r in df.iterrows():
        code = str(r.get("CodeElement", "")).strip()
        typ = str(r.get("NatureElement", "")).strip()
        name = str(r.get("CatalogName", "")).strip()
        pol = str(r.get("DetectedPolarity", "")).strip() or str(r.get("Polaritate", "")).strip()

        key = (typ, code, pol)
        if key in seen:
            continue
        seen.add(key)

        detected_sev = str(r.get("DetectedSeverity", "")).strip()
        lex_intensity = str(r.get("Intensitate", "")).strip()
        severity_score = str(r.get("SeverityScore", "")).strip()

        if detected_sev:
            estimated_intensity = detected_sev
        elif lex_intensity:
            estimated_intensity = lex_intensity
        else:
            estimated_intensity = severity_label_from_score(severity_score)

        context_parts = []
        for c in ["TriggerContext", "BodyRegion", "ClinicalIntent"]:
            v = str(r.get(c, "")).strip()
            if v and v not in context_parts:
                context_parts.append(v)

        temporal_parts = []
        for c in ["DetectedTemporality", "TemporalContext"]:
            v = str(r.get(c, "")).strip()
            if v and v not in temporal_parts:
                temporal_parts.append(v)

        rows.append({
            "Type": typ,
            "Code": code,
            "Element": name,
            "Polarity": pol,
            "EstimatedIntensity": estimated_intensity,
            "TemporalContext": "; ".join(temporal_parts),
            "ClinicalContext": "; ".join(context_parts),
            "MatchedExpression": str(r.get("matched_expression", "")).strip(),
        })

    out = pd.DataFrame(rows)

    if out.empty:
        return out

    sort_order = {"Sympt": 1, "Signe": 2, "RiskF": 3}
    out["_sort"] = out["Type"].map(sort_order).fillna(9)
    out = out.sort_values(["_sort", "Polarity", "Code"]).drop(columns=["_sort"]).reset_index(drop=True)
    return out


def safe_sheet(sheets, name):
    return sheets.get(name, pd.DataFrame())


def save_sheet_to_workbook(workbook_path, sheet_name, new_df):
    """
    Replace a single sheet inside the workbook while preserving the other sheets.
    Optimizat: foloseste openpyxl direct — nu reciteste/rescrie intregul workbook.
    Actualizeaza session_state in-place fara reread de pe disc.
    """
    from openpyxl import load_workbook as _openpyxl_lw

    workbook_path = Path(workbook_path)
    df_clean = new_df.fillna("")

    # ── Scrie selectiv pe disc ────────────────────────────────────────────
    wb = _openpyxl_lw(workbook_path)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.append(list(df_clean.columns))
    for row in df_clean.itertuples(index=False, name=None):
        ws.append([str(v) if v is not None else "" for v in row])

    # Reordoneaza sheet-urile
    preferred_order = ["Lexicon", "Dialogues", "Annotations", "PatientVectors",
                       "MedDiagInput", "ClinicalSynthesis", "AutoSuggestions",
                       "Unmatched", "Summary", "LoaderLog"]
    current = wb.sheetnames
    ordered = [s for s in preferred_order if s in current] +               [s for s in current if s not in preferred_order]
    wb._sheets = [wb[s] for s in ordered if s in current]
    wb.save(workbook_path)

    # ── Actualizeaza session_state in-place — fara reread ────────────────
    _wb_update_sheet(str(workbook_path), sheet_name, df_clean)


st.title("🩺 NlpRO")
st.caption("Romanian Clinical NLP Corpus • Anamneză → Annotations → Patient Vector → MedDiag Bridge → Clinical Synthesis")

with st.sidebar:
    st.header("Configuration")

    workbook_path_str = st.text_input("Workbook Excel", value=str(DEFAULT_WORKBOOK))
    pipeline_script_str = st.text_input("Pipeline script", value=str(DEFAULT_PIPELINE_SCRIPT))
    tabel2_path_str = st.text_input(
        "Tabel2 TITUS",
        value=str(BASE_DIR / "Tabel2_Titus_NumeElement.xlsx"),
    )

    workbook_path = Path(workbook_path_str)
    pipeline_script = Path(pipeline_script_str)

    st.divider()

    if st.button("🔄 Reload workbook"):
        _wb_invalidate(str(workbook_path))
        st.rerun()

    if st.button("▶️ Run pipeline"):
        try:
            code, stdout, stderr = run_pipeline(pipeline_script)
            if code == 0:
                st.success("Pipeline rulat cu succes.")
            else:
                st.error(f"Pipeline error code: {code}")

            with st.expander("Pipeline output"):
                st.code(stdout or "(no stdout)")
                if stderr:
                    st.code(stderr)
        except Exception as exc:
            st.error(str(exc))


if not workbook_path.exists():
    st.error(f"Workbook-ul nu există: {workbook_path}")
    st.stop()

sheets = _wb_get(str(workbook_path))
if not sheets:
    st.error("Workbook-ul nu a putut fi încărcat.")
    st.stop()

lexicon = safe_sheet(sheets, "Lexicon")
dialogues = safe_sheet(sheets, "Dialogues")
annotations = safe_sheet(sheets, "Annotations")
patient_vectors = safe_sheet(sheets, "PatientVectors")
mediag = safe_sheet(sheets, "MedDiagInput")
synthesis = safe_sheet(sheets, "ClinicalSynthesis")
unmatched = safe_sheet(sheets, "Unmatched")
autosuggestions = safe_sheet(sheets, "AutoSuggestions")
summary = safe_sheet(sheets, "Summary")
runtime_audit = safe_sheet(sheets, "LexiconRuntimeAudit")
dialog_match_audit = safe_sheet(sheets, "LexiconDialogMatchAudit")

# Status is hidden by default to keep the interface clinically focused.
with st.expander("Advanced / Debug status", expanded=False):
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Lexicon rows", len(lexicon))
    c2.metric("Dialogues turns", len(dialogues))
    c3.metric("Annotations", len(annotations))
    c4.metric("Patient vectors", len(patient_vectors))
    c5.metric("Unmatched", len(unmatched))

    if not annotations.empty and "DetectedPolarity" in annotations.columns:
        p1, p2, p3 = st.columns(3)
        p1.metric("Present", int((annotations["DetectedPolarity"] == "prezent").sum()))
        p2.metric("Absent", int((annotations["DetectedPolarity"] == "absent").sum()))
        p3.metric("Resolved", int((annotations["DetectedPolarity"] == "rezolvat").sum()))

st.divider()
st.subheader("Consultație activă")

if not dialogues.empty and "dialog_id" in dialogues.columns:
    all_dialog_ids = sorted(dialogues["dialog_id"].astype(str).dropna().unique())
else:
    all_dialog_ids = []

if "active_dialog_id" not in st.session_state:
    st.session_state["active_dialog_id"] = all_dialog_ids[0] if all_dialog_ids else ""

active_dialog_id = st.selectbox(
    "Select active dialog_id",
    all_dialog_ids,
    index=all_dialog_ids.index(st.session_state["active_dialog_id"]) if st.session_state["active_dialog_id"] in all_dialog_ids else 0,
    key="active_dialog_id",
)

def filter_by_active_dialog(df, dialog_id):
    if df is None or df.empty or "dialog_id" not in df.columns or not dialog_id:
        return df.iloc[0:0].copy() if isinstance(df, pd.DataFrame) else pd.DataFrame()
    return df[df["dialog_id"].astype(str) == str(dialog_id)].copy()

active_dialogues = filter_by_active_dialog(dialogues, active_dialog_id)
active_annotations = filter_by_active_dialog(annotations, active_dialog_id)
active_patient_vectors = filter_by_active_dialog(patient_vectors, active_dialog_id)
active_mediag = filter_by_active_dialog(mediag, active_dialog_id)
active_synthesis = filter_by_active_dialog(synthesis, active_dialog_id)
active_unmatched = filter_by_active_dialog(unmatched, active_dialog_id)
active_autosuggestions = filter_by_active_dialog(autosuggestions, active_dialog_id)

with st.expander("Detalii tehnice pentru consultația activă", expanded=False):
    a1, a2, a3, a4 = st.columns(4)
    a1.metric("Active turns", len(active_dialogues))
    a2.metric("Active annotations", len(active_annotations))
    a3.metric("Active patient vector rows", len(active_patient_vectors))
    a4.metric("Active unmatched", len(active_unmatched))
    st.metric("Active autosuggestions", len(active_autosuggestions))

tabs = st.tabs([
    "ClinicalSynthesis",
    "MedDiagInput",
    "TITUS Diagnostic",
    "Annotations",
    "Transcript",
    "Load transcript",
    "AutoSuggestions",
    "Unmatched",
    "Lexicon",
    "Correct Lexicon",
    "Runtime Audit",
    "Advanced",
])


with tabs[0]:
    st.subheader(f"ClinicalSynthesis — {active_dialog_id}")

    if not active_synthesis.empty:
        row = active_synthesis.head(1)
        r = row.iloc[0]
        st.markdown("### Sinteză clinică")
        st.write(r.get("ClinicalSynthesis", ""))
        st.markdown("### Orientare clinică")
        st.write(r.get("ClinicalOrientation", ""))
        st.markdown("### Triage hint")
        st.warning(r.get("TriageHint", ""))
        with st.expander("Rând complet ClinicalSynthesis", expanded=False):
            st.dataframe(row, use_container_width=True)
    else:
        st.info("Nu există sinteză clinică pentru consultația selectată. Rulează pipeline-ul.")


with tabs[1]:
    st.subheader(f"MedDiagInput / DiagMed input — {active_dialog_id}")
    st.caption("Acesta este inputul care trebuie trimis către DiagMed pentru consultația selectată.")

    if not active_mediag.empty and "MedDiagPayloadJSON" in active_mediag.columns:
        row = active_mediag.head(1)
        payload = row.iloc[0].get("MedDiagPayloadJSON", "")

        try:
            parsed_payload = json.loads(payload)
        except Exception:
            parsed_payload = None

        display_table = build_mediag_display_table(active_annotations)

        st.markdown("### Elemente trimise către DiagMed")
        if not display_table.empty:
            st.dataframe(display_table, use_container_width=True, height=350)

            csv_payload = display_table.to_csv(index=False).encode("utf-8")
            st.download_button(
                label="Download selected DiagMed readable table CSV",
                data=csv_payload,
                file_name=f"DiagMed_elements_{active_dialog_id}.csv",
                mime="text/csv",
            )
        else:
            st.info("Nu există elemente anotate pentru consultația selectată.")

        if parsed_payload:
            with st.expander("Coduri brute DiagMed", expanded=False):
                c1, c2, c3, c4, c5 = st.columns(5)
                c1.write("Sympt")
                c1.code(",".join(parsed_payload.get("sympt", [])) or "—")
                c2.write("Signe")
                c2.code(",".join(parsed_payload.get("signe", [])) or "—")
                c3.write("RiskF")
                c3.code(",".join(parsed_payload.get("riskf", [])) or "—")
                c4.write("Absent")
                c4.code(",".join(parsed_payload.get("absent", [])) or "—")
                c5.write("Resolved")
                c5.code(",".join(parsed_payload.get("resolved", [])) or "—")

            with st.expander("Payload JSON complet", expanded=False):
                st.json(parsed_payload)
        else:
            st.code(payload)

        st.download_button(
            label="Download selected DiagMed payload JSON",
            data=payload.encode("utf-8"),
            file_name=f"DiagMed_payload_{active_dialog_id}.json",
            mime="application/json",
        )

        with st.expander("Rând complet MedDiagInput", expanded=False):
            st.dataframe(active_mediag, use_container_width=True, height=250)
    else:
        st.info("Nu există MedDiagInput pentru consultația selectată. Rulează pipeline-ul.")



with tabs[2]:
    render_titus_tab(active_mediag, tabel2_path_str, active_dialog_id)

with tabs[3]:
    st.subheader(f"Annotations — {active_dialog_id}")

    if not active_annotations.empty:
        preferred_cols = [
            "matched_expression",
            "CodeElement",
            "NatureElement",
            "CatalogName",
            "DetectedPolarity",
            "DetectedTemporality",
            "DetectedSeverity",
            "utterance",
        ]
        cols = [c for c in preferred_cols if c in active_annotations.columns]
        compact = active_annotations[cols] if cols else active_annotations
        st.dataframe(compact, use_container_width=True, height=450)

        with st.expander("Annotations complete", expanded=False):
            st.dataframe(active_annotations, use_container_width=True, height=550)
    else:
        st.info("Nu există annotations pentru consultația selectată. Verifică dacă pipeline-ul a fost rulat după import și dacă transcriptul conține expresii acoperite de Lexicon.")


with tabs[4]:
    st.subheader(f"Transcript — {active_dialog_id}")
    if not active_dialogues.empty:
        for _, r in active_dialogues.sort_values("turn_id").iterrows():
            speaker = str(r.get("speaker", "")).upper()
            text_value = str(r.get("text", ""))
            if speaker == "PATIENT":
                st.markdown(f"**PACIENT:** {text_value}")
            elif speaker == "DOCTOR":
                st.markdown(f"**MEDIC:** {text_value}")
            else:
                st.markdown(f"**{speaker}:** {text_value}")

        with st.expander("Tabel Dialogues", expanded=False):
            st.dataframe(active_dialogues, use_container_width=True, height=400)
    else:
        st.info("Nu există transcript pentru consultația selectată.")


with tabs[5]:
    st.subheader("Load consultation transcript")
    st.markdown("""
Format acceptat:

```text
CASE: Chest pain case
DIALOG_ID: DLG_TEST_001

Doctor: ...
Patient: ...
Doctor: ...
Patient: ...
```
""")

    uploaded = st.file_uploader("Încarcă transcript .txt", type=["txt"])
    manual_text = st.text_area(
        "Sau lipește transcriptul aici",
        height=300,
        placeholder="DIALOG_ID: DLG_TEST_001\n\nDoctor: ...\nPatient: ...",
    )

    replace_existing = st.checkbox("Înlocuiește dialogul dacă dialog_id există deja", value=True)

    if st.button("📥 Load transcript into workbook"):
        try:
            if uploaded is not None:
                text_value = uploaded.read().decode("utf-8")
            else:
                text_value = manual_text

            if not text_value.strip():
                st.warning("Nu există transcript de încărcat.")
            else:
                dialog_id, case_title, parsed = parse_transcript_text(text_value)

                if parsed.empty:
                    st.error("Nu am găsit replici Doctor:/Patient:.")
                else:
                    removed, loaded = append_dialogue_to_workbook(
                        workbook_path=workbook_path,
                        new_dialogue=parsed,
                        dialog_id=dialog_id,
                        case_title=case_title,
                        replace_existing=replace_existing,
                    )

                    st.success(f"Dialog încărcat: {dialog_id} | replici: {loaded} | rânduri vechi eliminate: {removed}")

                    with st.spinner("Rulez analiza NLP pentru noul transcript..."):
                        code, stdout, stderr = run_pipeline(pipeline_script)

                    if code == 0:
                        st.success("Analiza NLP a fost rulată automat.")
                        st.session_state["active_dialog_id"] = dialog_id
                        _wb_invalidate(str(workbook_path))
                        st.rerun()
                    else:
                        st.error(f"Pipeline error code: {code}")
                        with st.expander("Pipeline output"):
                            st.code(stdout or "(no stdout)")
                            if stderr:
                                st.code(stderr)

                    st.dataframe(parsed, use_container_width=True)
        except Exception as exc:
            st.error(str(exc))


with tabs[6]:
    st.subheader(f"AutoSuggestions — {active_dialog_id}")
    st.caption("Validează sugestiile aici. Setează `ValidationStatus` la `accept` sau `reject`, apoi apasă Save.")

    show_all_auto = st.checkbox("Show all autosuggestions, not only active dialogue", value=False)
    auto_view = autosuggestions if show_all_auto else active_autosuggestions

    if not auto_view.empty:
        editable_cols = [
            "dialog_id",
            "turn_id",
            "utterance",
            "CandidateExpression",
            "AlreadyMatchedCodes",
            "SuggestedCodeElement",
            "SuggestedNatureElement",
            "SuggestedCatalogName",
            "SuggestedElementStandard",
            "SuggestionMethod",
            "SuggestionConfidence",
            "SuggestionReason",
            "ValidationStatus",
            "ProposedElementStandard",
            "ProposedCodeElement",
            "ProposedNatureElement",
            "ProposedCatalogName",
            "SemanticDomain",
        ]
        editable_cols = [c for c in editable_cols if c in auto_view.columns]
        edit_df = auto_view[editable_cols].copy()

        st.markdown("### Validation editor")
        edited = st.data_editor(
            edit_df,
            use_container_width=True,
            height=520,
            num_rows="fixed",
            column_config={
                "ValidationStatus": st.column_config.SelectboxColumn(
                    "ValidationStatus",
                    help="Alege accept pentru integrare în Lexicon, reject pentru respingere, suggested pentru nevalidat.",
                    options=["suggested", "accept", "reject", "edit"],
                    required=False,
                ),
                "SuggestedCodeElement": st.column_config.TextColumn("SuggestedCodeElement"),
                "SuggestedNatureElement": st.column_config.SelectboxColumn(
                    "SuggestedNatureElement",
                    options=["Sympt", "Signe", "RiskF", ""],
                    required=False,
                ),
                "SuggestedCatalogName": st.column_config.TextColumn("SuggestedCatalogName"),
                "ProposedCodeElement": st.column_config.TextColumn(
                    "ProposedCodeElement",
                    help="Completează aici dacă modifici sugestia.",
                ),
                "ProposedNatureElement": st.column_config.SelectboxColumn(
                    "ProposedNatureElement",
                    options=["", "Sympt", "Signe", "RiskF"],
                    required=False,
                ),
                "ProposedCatalogName": st.column_config.TextColumn("ProposedCatalogName"),
            },
            disabled=[
                c for c in editable_cols
                if c not in {
                    "ValidationStatus",
                    "ProposedElementStandard",
                    "ProposedCodeElement",
                    "ProposedNatureElement",
                    "ProposedCatalogName",
                    "SemanticDomain",
                }
            ],
            key=f"autosuggestions_editor_{active_dialog_id}_{show_all_auto}",
        )

        col_save, col_run = st.columns(2)

        with col_save:
            if st.button("💾 Save AutoSuggestions validation"):
                try:
                    full_auto = autosuggestions.copy()

                    # Preserve all columns and update only displayed rows.
                    key_cols = ["dialog_id", "turn_id", "CandidateExpression", "SuggestedCodeElement"]
                    for c in key_cols:
                        if c not in full_auto.columns:
                            full_auto[c] = ""
                        if c not in edited.columns:
                            edited[c] = ""

                    editable_update_cols = [
                        "ValidationStatus",
                        "ProposedElementStandard",
                        "ProposedCodeElement",
                        "ProposedNatureElement",
                        "ProposedCatalogName",
                        "SemanticDomain",
                    ]

                    for _, erow in edited.iterrows():
                        mask = (
                            (full_auto["dialog_id"].astype(str) == str(erow.get("dialog_id", "")))
                            & (full_auto["turn_id"].astype(str) == str(erow.get("turn_id", "")))
                            & (full_auto["CandidateExpression"].astype(str) == str(erow.get("CandidateExpression", "")))
                            & (full_auto["SuggestedCodeElement"].astype(str) == str(erow.get("SuggestedCodeElement", "")))
                        )

                        for col in editable_update_cols:
                            if col in full_auto.columns and col in edited.columns:
                                full_auto.loc[mask, col] = erow.get(col, "")

                    save_sheet_to_workbook(workbook_path, "AutoSuggestions", full_auto)
                    st.success("Validările au fost salvate în foaia AutoSuggestions.")
                    st.info("Acum rulează pipeline-ul pentru integrarea rândurilor cu ValidationStatus = accept.")
                    st.rerun()

                except Exception as exc:
                    st.error(f"Nu am putut salva validările: {exc}")

        with col_run:
            if st.button("▶️ Save + Run pipeline"):
                try:
                    full_auto = autosuggestions.copy()
                    key_cols = ["dialog_id", "turn_id", "CandidateExpression", "SuggestedCodeElement"]
                    for c in key_cols:
                        if c not in full_auto.columns:
                            full_auto[c] = ""
                        if c not in edited.columns:
                            edited[c] = ""

                    editable_update_cols = [
                        "ValidationStatus",
                        "ProposedElementStandard",
                        "ProposedCodeElement",
                        "ProposedNatureElement",
                        "ProposedCatalogName",
                        "SemanticDomain",
                    ]

                    for _, erow in edited.iterrows():
                        mask = (
                            (full_auto["dialog_id"].astype(str) == str(erow.get("dialog_id", "")))
                            & (full_auto["turn_id"].astype(str) == str(erow.get("turn_id", "")))
                            & (full_auto["CandidateExpression"].astype(str) == str(erow.get("CandidateExpression", "")))
                            & (full_auto["SuggestedCodeElement"].astype(str) == str(erow.get("SuggestedCodeElement", "")))
                        )

                        for col in editable_update_cols:
                            if col in full_auto.columns and col in edited.columns:
                                full_auto.loc[mask, col] = erow.get(col, "")

                    save_sheet_to_workbook(workbook_path, "AutoSuggestions", full_auto)

                    with st.spinner("Rulez pipeline-ul după validare..."):
                        code, stdout, stderr = run_pipeline(pipeline_script)

                    if code == 0:
                        st.success("Pipeline rulat. Sugestiile acceptate au fost integrate în Lexicon.")
                        _wb_invalidate(str(workbook_path))
                        st.rerun()
                    else:
                        st.error(f"Pipeline error code: {code}")
                        with st.expander("Pipeline output"):
                            st.code(stdout or "(no stdout)")
                            if stderr:
                                st.code(stderr)

                except Exception as exc:
                    st.error(f"Eroare la Save + Run: {exc}")

        with st.expander("AutoSuggestions complet", expanded=False):
            st.dataframe(auto_view, use_container_width=True, height=550)

        dataframe_download_button(
            auto_view,
            f"NlpRO_AutoSuggestions_{active_dialog_id if not show_all_auto else 'ALL'}.xlsx",
            "Download displayed AutoSuggestions as Excel",
        )
    else:
        st.info("Nu există autosuggestions pentru consultația selectată.")


with tabs[7]:
    st.subheader(f"Unmatched — {active_dialog_id}")
    st.markdown("Completează în Excel: `CandidateExpression`, `ProposedCodeElement`, `ProposedNatureElement`, `ProposedCatalogName`.")

    show_all_unmatched = st.checkbox("Show all unmatched, not only active dialogue", value=False)
    unmatched_view = unmatched if show_all_unmatched else active_unmatched

    st.dataframe(unmatched_view, use_container_width=True, height=550)
    if not unmatched_view.empty:
        dataframe_download_button(
            unmatched_view,
            f"NlpRO_Unmatched_{active_dialog_id if not show_all_unmatched else 'ALL'}.xlsx",
            "Download displayed Unmatched as Excel",
        )


with tabs[8]:
    st.subheader("Lexicon")
    if not lexicon.empty:
        search = st.text_input("Search lexicon", "")
        view = lexicon
        if search.strip():
            mask = view.astype(str).apply(lambda col: col.str.contains(search, case=False, na=False)).any(axis=1)
            view = view[mask]
        st.dataframe(view, use_container_width=True, height=550)
    else:
        st.info("No lexicon.")



with tabs[9]:
    st.subheader("Correct Lexicon")
    st.caption("Corectează un termen deja validat și intrat în Lexicon. Modificarea afectează rulările viitoare.")

    if not lexicon.empty:
        search_term = st.text_input("Caută expresie / cod / nume", "", key="lexicon_correction_search")

        lex_view = lexicon.copy()
        if search_term.strip():
            mask = lex_view.astype(str).apply(
                lambda col: col.str.contains(search_term, case=False, na=False)
            ).any(axis=1)
            lex_view = lex_view[mask]

        editable_cols = [
            "ExprID",
            "ExpresiePacient",
            "CodeElement",
            "Nature Element",
            "TypeElement",
            "CatalogName",
            "ElementStandard",
            "SemanticDomain",
            "Intensitate",
            "Polaritate",
            "ReviewStatus",
            "ReviewerNote",
        ]
        editable_cols = [c for c in editable_cols if c in lex_view.columns]

        edited_lex = st.data_editor(
            lex_view[editable_cols],
            use_container_width=True,
            height=520,
            num_rows="fixed",
            disabled=[c for c in editable_cols if c == "ExprID"],
            key="lexicon_correction_editor",
        )

        def _apply_lexicon_editor_changes(full_lex, edited):
            if "ExprID" not in full_lex.columns or "ExprID" not in edited.columns:
                raise ValueError("Lipsește coloana ExprID; nu pot identifica rândurile din Lexicon.")

            full_lex = full_lex.copy()

            for _, erow in edited.iterrows():
                expr_id = str(erow.get("ExprID", "")).strip()
                if not expr_id:
                    continue

                mask = full_lex["ExprID"].astype(str) == expr_id

                for col in edited.columns:
                    if col in full_lex.columns and col != "ExprID":
                        full_lex.loc[mask, col] = erow.get(col, "")

            return full_lex

        c_save, c_run = st.columns(2)

        with c_save:
            if st.button("💾 Save Lexicon corrections"):
                try:
                    updated_lex = _apply_lexicon_editor_changes(lexicon, edited_lex)
                    save_sheet_to_workbook(workbook_path, "Lexicon", updated_lex)
                    st.success("Corecțiile din Lexicon au fost salvate.")
                    st.info("Rulează pipeline-ul pentru a recalcula Annotations / MedDiagInput.")
                    st.rerun()
                except Exception as exc:
                    st.error(f"Eroare la salvarea Lexicon: {exc}")

        with c_run:
            if st.button("▶️ Save Lexicon + Run pipeline"):
                try:
                    updated_lex = _apply_lexicon_editor_changes(lexicon, edited_lex)
                    save_sheet_to_workbook(workbook_path, "Lexicon", updated_lex)

                    with st.spinner("Rulez pipeline-ul după corecția Lexicon..."):
                        code, stdout, stderr = run_pipeline(pipeline_script)

                    if code == 0:
                        st.success("Pipeline rulat după corecția Lexicon.")
                        _wb_invalidate(str(workbook_path))
                        st.rerun()
                    else:
                        st.error(f"Pipeline error code: {code}")
                        with st.expander("Pipeline output"):
                            st.code(stdout or "(no stdout)")
                            if stderr:
                                st.code(stderr)
                except Exception as exc:
                    st.error(f"Eroare la Save + Run: {exc}")
    else:
        st.info("Lexicon gol sau neîncărcat.")



with tabs[10]:
    st.subheader("Runtime Audit")
    st.caption("Diagnostic pentru rândurile din Lexicon: intră în runtime sau sunt excluse? Se potrivesc textual în vreun dialog?")

    search_audit = st.text_input("Caută în audit: expresie / cod / dialog_id", "", key="runtime_audit_search")

    audit_view = runtime_audit.copy()
    if search_audit.strip() and not audit_view.empty:
        mask = audit_view.astype(str).apply(lambda col: col.str.contains(search_audit, case=False, na=False)).any(axis=1)
        audit_view = audit_view[mask]

    st.markdown("### LexiconRuntimeAudit")
    st.dataframe(audit_view, use_container_width=True, height=420)

    st.markdown("### LexiconDialogMatchAudit")
    match_view = dialog_match_audit.copy()
    if search_audit.strip() and not match_view.empty:
        mask = match_view.astype(str).apply(lambda col: col.str.contains(search_audit, case=False, na=False)).any(axis=1)
        match_view = match_view[mask]
    st.dataframe(match_view, use_container_width=True, height=320)


with tabs[11]:
    st.subheader("Advanced")

    st.markdown("### PatientVectors")
    st.dataframe(active_patient_vectors, use_container_width=True, height=300)

    st.markdown("### Summary")
    st.dataframe(summary, use_container_width=True, height=250)

    st.markdown("### Downloads")

    if workbook_path.exists():
        st.download_button(
            label="Download current workbook",
            data=workbook_path.read_bytes(),
            file_name=workbook_path.name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    if pipeline_script.exists():
        st.download_button(
            label="Download pipeline script",
            data=pipeline_script.read_bytes(),
            file_name=pipeline_script.name,
            mime="text/x-python",
        )